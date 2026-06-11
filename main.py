from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse
from typing import List
import openpyxl
import hashlib
import io
import re
import uuid
import asyncio
import urllib.request
import json
import os
from datetime import datetime, date, timedelta

import time

app = FastAPI(title="Sheet Discovery Service", version="3.1.0")

JOB_TIMEOUT_SECONDS     = 120  # 2 minutes per job
SESSION_TIMEOUT_SECONDS = 600  # 10 minutes per session

@app.on_event("startup")
async def startup():
    asyncio.create_task(cleanup_stale_jobs())

async def cleanup_stale_jobs():
    while True:
        await asyncio.sleep(60)  # check every minute
        now = time.time()

        # Find stale jobs
        stale_job_ids = [
            jid for jid, j in list(_jobs.items())
            if now - j.get("created_at", now) > JOB_TIMEOUT_SECONDS
        ]

        sessions_to_advance = set()
        for jid in stale_job_ids:
            job = _jobs.pop(jid, {})
            sid = job.get("session_id")
            if sid and sid in _sessions:
                session = _sessions[sid]
                session["_pending_jobs"].discard(jid)
                session.setdefault("errors", []).append(
                    f"Job {jid} (stage: {job.get('stage')}) timed out after {JOB_TIMEOUT_SECONDS}s"
                )
                # If no more pending jobs, this session needs to advance
                if not session["_pending_jobs"]:
                    sessions_to_advance.add((sid, job.get("stage")))

        # Advance sessions whose last pending job timed out
        for sid, stage in sessions_to_advance:
            if sid not in _sessions:
                continue
            session = _sessions[sid]
            if stage == "qualify":
                asyncio.create_task(advance_from_qualify(sid))
            elif stage == "postgres_sku":
                session["postgres_results"] = {"matches": [], "error": "timed out"}
                asyncio.create_task(advance_from_postgres(sid))
            elif stage == "classify_sheet":
                asyncio.create_task(stage_identify_retailer(sid))
            elif stage == "identify_retailer":
                session["retailer"] = None
                session["flags"]["retailer_identification"] = "timed out"
                asyncio.create_task(stage_date_config(sid))
            elif stage == "date_config":
                asyncio.create_task(stage_multisheet(sid))
            elif stage == "write_audit":
                asyncio.create_task(stage_assemble(sid))
            elif stage == "write_config":
                asyncio.create_task(stage_assemble(sid))

        # Find stale sessions
        stale_session_ids = [
            sid for sid, s in list(_sessions.items())
            if now - s.get("created_at", now) > SESSION_TIMEOUT_SECONDS
            and s.get("status") not in ("complete", "failed")
        ]

        for sid in stale_session_ids:
            session = _sessions.get(sid, {})
            session["stage"]  = "failed"
            session["status"] = "failed"
            session["result"] = {
                "error": "Session timed out",
                "errors": session.get("errors", []),
            }

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
N8N_AI_WEBHOOK       = os.environ.get("N8N_AI_WEBHOOK",       "http://n8n:5678/webhook/ai")
N8N_POSTGRES_WEBHOOK = os.environ.get("N8N_POSTGRES_WEBHOOK", "http://n8n:5678/webhook/postgres")

# ─────────────────────────────────────────────
# SESSION STORE
# ─────────────────────────────────────────────
_sessions: dict = {}  # session_id -> session
_jobs: dict     = {}  # job_id -> {session_id, stage, key, ...}


# ─────────────────────────────────────────────
# WEBHOOK HELPER
# ─────────────────────────────────────────────

def fire_webhook(url: str, job_id: str, payload: dict) -> str | None:
    body = json.dumps({"job_id": job_id, **payload}).encode()
    try:
        req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
        urllib.request.urlopen(req, timeout=5)
        return None
    except Exception as e:
        return str(e)


# ─────────────────────────────────────────────
# CELL CLASSIFICATION
# ─────────────────────────────────────────────

DATE_RANGE_RE  = re.compile(r"^\d{1,2}/\d{1,2}[-–]\d{1,2}/\d{1,2}$")
FISCAL_WEEK_RE = re.compile(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+wk\s*\d+$", re.IGNORECASE)
WK_NUMBER_RE   = re.compile(r"^wk\s*\d+$", re.IGNORECASE)
YEAR_RE        = re.compile(r"\b(20\d{2})\b")

def classify_cell(val) -> str:
    if val is None:               return "empty"
    if isinstance(val, bool):     return "bool"
    if isinstance(val, datetime): return "datetime"
    if isinstance(val, int):      return "integer"
    if isinstance(val, float):
        return "float_0_to_1" if 0 <= val <= 1 else "float_above_1"
    if isinstance(val, str):
        s = val.strip()
        if DATE_RANGE_RE.match(s):  return "date_range_string"
        if FISCAL_WEEK_RE.match(s): return "fiscal_week_label"
        if WK_NUMBER_RE.match(s):   return "week_number_label"
        return "string"
    return "unknown"

def is_date_like(val) -> bool:
    return classify_cell(val) in (
        "datetime", "date_range_string", "fiscal_week_label", "week_number_label"
    )


# ─────────────────────────────────────────────
# GRID DETECTION
# ─────────────────────────────────────────────

def find_date_axis(rows) -> dict | None:
    best = {"row": None, "count": 0, "cols": [], "samples": [], "format": None}
    for row_idx, row in enumerate(rows[:15]):
        date_cols = [(ci, v) for ci, v in enumerate(row) if is_date_like(v)]
        if len(date_cols) > best["count"]:
            formats = set(classify_cell(v) for _, v in date_cols)
            best = {
                "row": row_idx, "count": len(date_cols),
                "cols": [ci for ci, _ in date_cols],
                "samples": [str(v) for _, v in date_cols[:8]],
                "format": list(formats)[0] if len(formats) == 1 else "mixed",
            }
    if best["count"] < 2:
        return None

    cols = best["cols"]
    year_present = best["format"] == "datetime" or any(YEAR_RE.search(s) for s in best["samples"])

    interleaved = False
    if len(cols) >= 3:
        gaps = [cols[i+1] - cols[i] for i in range(len(cols)-1)]
        interleaved = len(set(gaps)) == 1 and gaps[0] == 2

    year_boundary_detected = False
    if best["format"] in ("date_range_string", "mixed"):
        months = set()
        for s in best["samples"]:
            m = re.match(r'^(\d{1,2})/', s.strip())
            if m:
                months.add(int(m.group(1)))
        if 12 in months and 1 in months:
            year_boundary_detected = True

    return {
        "row": best["row"], "col_count": best["count"], "cols": cols,
        "sample_values": best["samples"], "format": best["format"],
        "year_present": year_present, "interleaved_empty_cols": interleaved,
        "year_boundary_detected": year_boundary_detected,
    }


def find_data_start_row(rows, date_axis_row: int, date_cols: list) -> int:
    for row_idx in range(date_axis_row + 1, min(date_axis_row + 10, len(rows))):
        row = rows[row_idx]
        vals = [row[dc] for dc in date_cols if dc < len(row) and row[dc] is not None]
        numeric = [v for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if numeric:
            return row_idx
    return date_axis_row + 1


def find_sku_candidates(rows, date_axis_row: int, date_cols: list) -> list:
    if not date_cols:
        return []
    left_boundary = min(date_cols)
    data_start = find_data_start_row(rows, date_axis_row, date_cols)
    data_rows = [rows[r] for r in range(data_start, min(data_start + 50, len(rows)))]
    total_data_rows = len(data_rows)
    candidates = []

    for col_idx in range(left_boundary):
        col_vals = [row[col_idx] for row in data_rows if col_idx < len(row) and row[col_idx] is not None]
        if not col_vals:
            continue

        types = [classify_cell(v) for v in col_vals]
        type_counts = {}
        for t in types:
            type_counts[t] = type_counts.get(t, 0) + 1
        dominant = max(type_counts, key=type_counts.get)
        fill_rate = round(len(col_vals) / total_data_rows, 2) if total_data_rows > 0 else 0

        pre_data_strings = [
            {"row": row_idx, "value": rows[row_idx][col_idx]}
            for row_idx in range(data_start)
            if col_idx < len(rows[row_idx])
            and isinstance(rows[row_idx][col_idx], str)
            and rows[row_idx][col_idx].strip()
        ]

        candidates.append({
            "col": col_idx,
            "pre_data_strings": pre_data_strings,
            "dominant_type": dominant,
            "type_distribution": type_counts,
            "fill_rate": fill_rate,
            "sample_values": [str(v) for v in col_vals[:5]],
        })
    return candidates


EMBEDDED_PATTERNS = [
    ("integer_dash_description",        re.compile(r'^(\d{4,10})\s*[-\u2013]\s*(.{3,})$')),
    ("alphanumeric_space_description",  re.compile(r'^([A-Z0-9]{2,}-[A-Z0-9\-]{2,})\s+(.{3,})$', re.IGNORECASE)),
    ("description_space_alphanumeric",  re.compile(r'^(.{3,})\s{2,}([A-Z0-9]{2,}-[A-Z0-9\-]{2,})\s*$', re.IGNORECASE)),
    ("description_space_nodash_sku",    re.compile(r'^(.{5,})\s+([A-Z]{2,}\d{3,})\s*$', re.IGNORECASE)),
]


def detect_embedded_sku(rows, date_cols: list, data_start_row: int) -> list:
    if not date_cols:
        return []
    left_boundary = min(date_cols)
    embedded_candidates = []

    for col_idx in range(left_boundary):
        col_vals = [
            str(rows[r][col_idx]).strip()
            for r in range(data_start_row, min(data_start_row + 30, len(rows)))
            if col_idx < len(rows[r]) and rows[r][col_idx] is not None
            and isinstance(rows[r][col_idx], str)
        ]
        if len(col_vals) < 3:
            continue

        col_matches = []
        seen = set()
        for pattern_name, pattern_re in EMBEDDED_PATTERNS:
            for val in col_vals:
                if val not in seen:
                    m = pattern_re.match(val)
                    if m:
                        seen.add(val)
                        # description_space_* patterns have description in group(1), SKU in group(2)
                        if pattern_name.startswith("description_space"):
                            sku, desc = m.group(2).strip(), m.group(1).strip()
                        else:
                            sku, desc = m.group(1).strip(), m.group(2).strip()
                        col_matches.append({
                            "raw": val, "sku": sku,
                            "description": desc, "pattern": pattern_name,
                        })
        if col_matches:
            embedded_candidates.append({
                "col": col_idx, "total_values": len(col_vals),
                "matched_values": len(col_matches), "extractions": col_matches,
            })
    return embedded_candidates


def find_year_anchors(filename: str, rows: list) -> list:
    anchors = []
    for m in YEAR_RE.finditer(filename):
        anchors.append({"source": "filename", "value": m.group(1)})
    for row_idx, row in enumerate(rows[:5]):
        for col_idx, val in enumerate(row):
            if isinstance(val, datetime):
                anchors.append({"source": "cell", "row": row_idx, "col": col_idx, "value": str(val.date())})
            elif isinstance(val, str):
                for m in YEAR_RE.finditer(val):
                    anchors.append({"source": "cell", "row": row_idx, "col": col_idx, "value": m.group(1)})
    return anchors[:6]


# ─────────────────────────────────────────────
# DATE NORMALIZATION — normalize to week-ending Saturday
# ─────────────────────────────────────────────

def normalize_to_saturday(d: date) -> date:
    """Return the Saturday of the week containing d (Mon=0, Sat=5)."""
    days_to_saturday = (5 - d.weekday()) % 7
    return d + timedelta(days=days_to_saturday)


# ─────────────────────────────────────────────
# QUALIFY SIGNALS
# ─────────────────────────────────────────────

def extract_qualify_signals(rows: list, sheet_name: str, filename: str) -> dict:
    integer_count = float_above_count = float_0_to_1_count = 0
    crosshair_sample = []

    date_axis = find_date_axis(rows)
    if date_axis:
        date_cols  = date_axis["cols"][:8]
        data_start = find_data_start_row(rows, date_axis["row"], date_axis["cols"])
        for row in rows[data_start:data_start + 50]:
            for dc in date_cols:
                if dc >= len(row) or row[dc] is None:
                    continue
                v = row[dc]
                if isinstance(v, bool):
                    continue
                if isinstance(v, int):
                    integer_count += 1
                    if len(crosshair_sample) < 40:
                        crosshair_sample.append(v)
                elif isinstance(v, float) and v > 1:
                    float_above_count += 1
                    if len(crosshair_sample) < 40:
                        crosshair_sample.append(v)
                elif isinstance(v, float) and 0 < v <= 1:
                    float_0_to_1_count += 1
                    if len(crosshair_sample) < 40:
                        crosshair_sample.append(v)
        total = integer_count + float_above_count + float_0_to_1_count
        if total == 0:                            dominant_type = None
        elif float_above_count / total > 0.7:     dominant_type = "float_above_1"
        elif float_0_to_1_count / total > 0.7:   dominant_type = "float_0_to_1"
        elif integer_count / total > 0.7:         dominant_type = "integer"
        else:                                     dominant_type = "mixed"
    else:
        for row in rows[:50]:
            for v in row:
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    if len(crosshair_sample) < 40:
                        crosshair_sample.append(v)
            if len(crosshair_sample) >= 40:
                break
        dominant_type = None

    column_labels = set()
    for row in rows[:6]:
        for val in row:
            if isinstance(val, str) and val.strip() and len(val.strip()) < 60:
                column_labels.add(val.strip())

    inline_strings = set()
    for row in rows[6:20]:
        for val in row:
            if isinstance(val, str) and val.strip() and len(val.strip()) < 80:
                inline_strings.add(val.strip())

    return {
        "sheet_name":       sheet_name,
        "filename":         filename,
        "dominant_type":    dominant_type,
        "crosshair_sample": crosshair_sample,
        "column_labels":    sorted(column_labels, key=len)[:10],
        "inline_strings":   sorted(inline_strings, key=len)[:10],
    }


def build_qualify_prompt(signals: dict) -> str:
    return f"""You are the gatekeeper of a retail sales data ingestion pipeline. A sheet has arrived and you must determine if it should be disqualified.

A sheet should be disqualified if it does NOT contain unit sales data. Unit sales data has integer values at the intersection of product identifiers and date columns.

Disqualify if any of the following are true:
- Crosshair values are decimals above 1 — dollar revenue
- Crosshair values are between 0 and 1 — percentage metrics
- Column labels contain $$$ or $$ — dollar sheet
- Sheet name contains CFP, Forecast, FCST, Projection, Order
- Vocabulary contains forecast or projection language

EVIDENCE:
Sheet name: {signals["sheet_name"]}
Filename: {signals["filename"]}
Crosshair sample values: {signals["crosshair_sample"]}
Dominant crosshair type: {signals["dominant_type"]}
Column labels: {signals["column_labels"]}
Inline strings: {signals["inline_strings"]}

Respond with JSON only:
{{"disqualified": true or false, "reason": "one sentence explanation"}}"""


# ─────────────────────────────────────────────
# POSTGRES SQL BUILDER
# ─────────────────────────────────────────────

def build_sku_lookup_sql(candidates: list) -> tuple[str, list]:
    """
    Build case-insensitive exact match query against inventory_view.
    Values embedded inline — n8n Postgres node does not support $1 params.
    Returns (sql, []) — empty params list.
    """
    # Safely quote each candidate for inline SQL
    quoted = ", ".join(
        "'" + str(c).replace("'", "''") + "'"
        for c in candidates
        if c and str(c).strip()
    )
    if not quoted:
        quoted = "''"

    sql = f"""SELECT inventory_sku, base_model, base_variant, description, upc
FROM inventory_view
WHERE UPPER(inventory_sku) = ANY(ARRAY[{quoted}]::text[])
   OR UPPER(base_variant)  = ANY(ARRAY[{quoted}]::text[])
   OR UPPER(base_model)    = ANY(ARRAY[{quoted}]::text[])"""

    return sql, []


# ─────────────────────────────────────────────
# COLUMN CLASSIFY PROMPT
# ─────────────────────────────────────────────

def build_column_classify_prompt(sheet_name: str, sku_candidates: list,
                                  embedded_sku: list, postgres_matches: list,
                                  col_candidates: dict) -> str:
    """
    Build a single prompt to classify all columns in a sheet.
    AI receives full context: columns, Postgres matches, embedded SKU detections.
    """
    matched_values = set()
    for row in postgres_matches:
        for field in ("inventory_sku", "base_variant", "base_model"):
            v = row.get(field, "")
            if v:
                matched_values.add(v.upper())

    cols_info = []
    for col in sku_candidates:
        col_idx = col["col"]
        label = col.get("pre_data_strings", [{}])[-1].get("value", "") if col.get("pre_data_strings") else ""
        samples = col.get("sample_values", [])
        postgres_confirmed = any(str(v).upper() in matched_values for v in samples)

        emb_matches = []
        for emb in embedded_sku:
            if emb["col"] == col_idx:
                for ext in emb.get("extractions", []):
                    sku = ext.get("sku", "")
                    if sku.upper() in matched_values:
                        emb_matches.append(sku)

        cols_info.append({
            "col":                  col_idx,
            "label":                label,
            "dominant_type":        col.get("dominant_type"),
            "sample_values":        samples,
            "postgres_matched":     postgres_confirmed,
            "embedded_sku_matches": emb_matches,
        })

    return f"""You are classifying columns in a retail sales spreadsheet sheet named "{sheet_name}".

For each column, classify it as one of:
- retailer_sku: the retailer's product identifier (WIC#, DPCI, Item#, UPC etc)
- supplier_sku: the supplier's internal SKU or style code
- description: product name or description text (may contain embedded supplier SKU)
- cost: unit cost or wholesale price
- retail_price: retail selling price
- inventory: stock quantities
- other: anything else

postgres_matched means the column's values were found in the supplier inventory database.
embedded_sku_matches lists supplier SKUs extracted from a description column that matched the inventory database.

COLUMNS:
{json.dumps(cols_info, indent=2)}

Respond with JSON only:
{{
  "columns": [
    {{"col": 0, "classification": "type", "confidence": "high/medium/low", "reason": "one sentence", "has_embedded_supplier_sku": true/false}},
    ...
  ]
}}"""


def build_classify_prompt(col: dict, confirmed_cols: list) -> str:
    confirmed_context = ""
    if confirmed_cols:
        confirmed_context = "\nAlready confirmed columns:\n"
        for c in confirmed_cols:
            confirmed_context += f"  col {c['col']} ({c['label']}) = {c['classification']}, samples: {c['sample_values'][:3]}\n"

    return f"""You are classifying a column in a retail sales spreadsheet.

Classify it as one of:
- retailer_sku: the retailer's product identifier (WIC#, DPCI, Item#, UPC etc)
- supplier_sku: the supplier's internal SKU or style code
- description: product name or description text
- cost: unit cost or wholesale price
- retail_price: retail selling price
- inventory: stock quantities
- other: anything else

COLUMN TO CLASSIFY:
Column index: {col["col"]}
Label: {col["label"]}
Dominant type: {col["dominant_type"]}
Sample values: {col["sample_values"]}
{confirmed_context}
Respond with JSON only:
{{"col": {col["col"]}, "classification": "type", "confidence": "high/medium/low", "reason": "one sentence"}}"""


# ─────────────────────────────────────────────
# DATE CONFIG PROMPT
# ─────────────────────────────────────────────

def build_date_prompt(date_axis: dict, year_anchors: list, sheet_name: str,
                       filename: str = "", cross_sheet_anchors: list = None) -> str:
    cross_context = ""
    if cross_sheet_anchors:
        cross_context = f"\nYear anchors from other sheets in same file: {cross_sheet_anchors}"

    return f"""You are configuring the date settings for a retail sales sheet.

Determine the year value and week convention from the evidence.
The file was received as: {filename}

Sheet: {sheet_name}
Date axis format: {date_axis["format"]}
Year present in dates: {date_axis["year_present"]}
Year boundary detected: {date_axis.get("year_boundary_detected", False)}
Sample date values: {date_axis["sample_values"]}
Year anchors in this sheet: {year_anchors}{cross_context}

Important: If dates span December and January, December dates belong to the EARLIER year and January dates belong to the LATER year. Use the file date and year anchors to determine which years those are.

All dates will be normalized to week-ending Saturday.

Respond with JSON only:
{{"year_value": 2026, "year_inference_strategy": "how year was determined", "week_convention": "what convention the source uses", "year_boundary_note": "null or explanation if dates span two years"}}"""



# ─────────────────────────────────────────────
# RETAILER IDENTIFICATION SQL
# ─────────────────────────────────────────────

def build_retailer_identify_sql(retailer_sku_candidates: list) -> str:
    """
    Query retailer_sku_map to identify retailer from SKU candidates.
    Returns retailer and match count — 3+ matches confirms identity.
    """
    quoted = ", ".join(
        "'" + str(c).replace("'", "''") + "'"
        for c in retailer_sku_candidates
        if c and str(c).strip()
    )
    if not quoted:
        quoted = "''"

    return f"""
SELECT retailer, COUNT(*) as matches
FROM retailer_sku_map
WHERE UPPER(retailer_sku) = ANY(ARRAY[{quoted}]::text[])
  AND active = true
GROUP BY retailer
ORDER BY matches DESC
LIMIT 1
""".strip()


def build_update_file_audit_sql(file_audit_id: str, discovery_result: dict,
                                 retailer: str | None, status: str) -> str:
    """
    Build SQL to update file_audit row with discovery result and retailer.
    """
    result_json = json.dumps(discovery_result).replace("'", "''")
    retailer_val = f"'{retailer}'" if retailer else "NULL"
    return f"""
UPDATE file_audit
SET discovery_result = '{result_json}'::jsonb,
    retailer         = {retailer_val},
    status           = '{status}',
    updated_at       = now()
WHERE id = '{file_audit_id}'
""".strip()


def build_insert_retailer_config_sql(retailer: str, file_audit_id: str,
                                      discovery_result: dict) -> str:
    """
    Insert a pending_review config row for a newly identified retailer.
    """
    def safe(v): return json.dumps(v).replace("'", "''")

    return f"""
INSERT INTO retailer_configs (
    retailer, status, version, file_audit_id,
    qualified_sheets, column_mapping, date_config, flags
)
VALUES (
    '{retailer}', 'pending_review', 1, '{file_audit_id}',
    '{safe(discovery_result.get("qualified_sheets", []))}'::jsonb,
    '{safe(discovery_result.get("column_mapping", {}))}'::jsonb,
    '{safe(discovery_result.get("date_config", {}))}'::jsonb,
    '{safe(discovery_result.get("flags", {}))}'::jsonb
)
ON CONFLICT DO NOTHING
""".strip()

# ─────────────────────────────────────────────
# PIPELINE STAGES
# ─────────────────────────────────────────────

async def stage_qualify(session_id: str):
    """Stage 1 — qualify each sheet."""
    session = _sessions[session_id]
    session["stage"] = "qualifying"
    session["status"] = "running"
    sheets = session["_sheets"]
    pending = 0

    for sheet_name, rows in sheets.items():
        signals = extract_qualify_signals(rows, sheet_name, session["filename"])

        # Deterministic
        if signals["dominant_type"] == "float_above_1":
            session["qualify_results"][sheet_name] = {
                "disqualified": True,
                "reason": "Crosshair values predominantly decimals above 1 — dollar revenue",
                "source": "python",
            }
            continue

        if signals["dominant_type"] == "float_0_to_1":
            session["qualify_results"][sheet_name] = {
                "disqualified": True,
                "reason": "Crosshair values predominantly between 0 and 1 — percentage metrics",
                "source": "python",
            }
            continue

        # Send to AI
        prompt = build_qualify_prompt(signals)
        job_id = str(uuid.uuid4())
        _jobs[job_id] = {
            "session_id": session_id,
            "stage":      "qualify",
            "sheet_name": sheet_name,
            "signals":    signals,
            "created_at": time.time(),
        }
        err = fire_webhook(N8N_AI_WEBHOOK, job_id, {"prompt": prompt})
        if err:
            session["qualify_results"][sheet_name] = {
                "disqualified": None,
                "reason": f"Webhook error: {err}",
                "source": "error",
            }
        else:
            session["status"] = "awaiting_ai"
            session["_pending_jobs"].add(job_id)
            pending += 1

    if pending == 0:
        await advance_from_qualify(session_id)


async def advance_from_qualify(session_id: str):
    """After all qualify jobs done — filter qualified sheets, advance to grid location."""
    session = _sessions[session_id]
    results = session["qualify_results"]

    qualified = [
        name for name, r in results.items()
        if not r.get("disqualified")
    ]

    session["qualified_sheets"] = qualified

    if not qualified:
        session["stage"] = "complete"
        session["status"] = "complete"
        session["result"] = {"status": "no_sales_data", "qualify_results": results}
        return

    await stage_locate_grid(session_id)


async def stage_locate_grid(session_id: str):
    """Stage 2 — locate grid for each qualified sheet. Pure Python."""
    session = _sessions[session_id]
    session["stage"] = "locating"
    session["status"] = "running"
    sheets = session["_sheets"]

    for sheet_name in session["qualified_sheets"]:
        rows = sheets[sheet_name]
        date_axis  = find_date_axis(rows)
        if not date_axis:
            session["grid"][sheet_name] = {"error": "No date axis found"}
            continue

        date_cols  = date_axis["cols"]
        data_start = find_data_start_row(rows, date_axis["row"], date_cols)
        candidates = find_sku_candidates(rows, date_axis["row"], date_cols)
        embedded   = detect_embedded_sku(rows, date_cols, data_start)
        anchors    = find_year_anchors(session["filename"], rows)

        session["grid"][sheet_name] = {
            "date_axis":      date_axis,
            "data_start_row": data_start,
            "sku_candidates": candidates,
            "embedded_sku":   embedded,
            "year_anchors":   anchors,
        }

    await stage_identify_columns(session_id)


async def stage_identify_columns(session_id: str):
    """Stage 3 — identify columns via Postgres then one AI call per sheet."""
    session = _sessions[session_id]
    session["stage"] = "identifying"
    session["status"] = "running"

    # Collect all unique values from every left-of-axis column, tagged with col index
    # No filtering — Python does not decide what is or isn't a SKU candidate
    col_candidates = {}  # value -> set of col indices it appears in

    for sheet_name in session["qualified_sheets"]:
        grid = session["grid"].get(sheet_name, {})

        for col in grid.get("sku_candidates", []):
            col_idx = col["col"]
            for v in col.get("sample_values", []):
                s = str(v).strip()
                if s:
                    col_candidates.setdefault(s, set()).add(col_idx)

        for emb in grid.get("embedded_sku", []):
            col_idx = emb["col"]
            for ext in emb.get("extractions", []):
                sku = ext.get("sku", "").strip()
                if sku:
                    col_candidates.setdefault(sku, set()).add(col_idx)

    # Store col_candidates for use after Postgres responds
    session["_col_candidates"] = {k: list(v) for k, v in col_candidates.items()}

    candidates = sorted(col_candidates.keys())
    sql, _ = build_sku_lookup_sql(candidates)

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {
        "session_id": session_id,
        "stage":      "postgres_sku",
        "created_at": time.time(),
    }

    err = fire_webhook(N8N_POSTGRES_WEBHOOK, job_id, {"sql": sql, "params": []})
    if err:
        session["postgres_results"] = {"matches": [], "matched_candidates": [], "error": err}
        await advance_from_postgres(session_id)
    else:
        session["status"] = "awaiting_postgres"
        session["_pending_jobs"].add(job_id)


async def advance_from_postgres(session_id: str):
    """After Postgres results — one AI call per sheet to classify all columns."""
    session = _sessions[session_id]
    pg      = session.get("postgres_results", {})

    for sheet_name in session["qualified_sheets"]:
        grid          = session["grid"].get(sheet_name, {})
        sku_candidates = grid.get("sku_candidates", [])
        embedded_sku   = grid.get("embedded_sku", [])
        matches        = pg.get("matches", [])

        prompt = build_column_classify_prompt(
            sheet_name, sku_candidates, embedded_sku, matches,
            session.get("_col_candidates", {})
        )

        job_id = str(uuid.uuid4())
        _jobs[job_id] = {
            "session_id": session_id,
            "stage":      "classify_sheet",
            "sheet_name": sheet_name,
            "created_at": time.time(),
        }

        err = fire_webhook(N8N_AI_WEBHOOK, job_id, {"prompt": prompt})
        if err:
            session["column_mapping"][sheet_name] = {"error": f"Webhook error: {err}"}
        else:
            session["status"] = "awaiting_ai"
            session["_pending_jobs"].add(job_id)

    if not session["_pending_jobs"]:
        await stage_identify_retailer(session_id)



async def stage_identify_retailer(session_id: str):
    """Stage 3b — identify retailer by querying retailer_sku_map."""
    session = _sessions[session_id]
    session["stage"] = "identifying_retailer"
    session["status"] = "running"

    # Collect ALL retailer SKU column values across all qualified sheets
    # Use full sheet data, not just samples — own Postgres, no cost concern
    retailer_sku_candidates = set()
    for sheet_name in session["qualified_sheets"]:
        mapping = session["column_mapping"].get(sheet_name, [])
        cols = mapping if isinstance(mapping, list) else mapping.get("columns", [])
        grid = session["grid"].get(sheet_name, {})
        rows = session["_sheets"].get(sheet_name, [])
        data_start = grid.get("data_start_row", 0)

        for col_info in cols:
            if col_info.get("classification") == "retailer_sku":
                col_idx = col_info.get("col")
                for row in rows[data_start:]:
                    if col_idx < len(row) and row[col_idx] is not None:
                        v = str(row[col_idx]).strip()
                        if v:
                            retailer_sku_candidates.add(v)

    if not retailer_sku_candidates:
        # No retailer SKU column found — skip, flag for review
        session["retailer"] = None
        session["flags"]["retailer_identification"] = "no_retailer_sku_column_found"
        await stage_date_config(session_id)
        return

    sql = build_retailer_identify_sql(sorted(retailer_sku_candidates))

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {
        "session_id": session_id,
        "stage":      "identify_retailer",
        "created_at": time.time(),
    }

    err = fire_webhook(N8N_POSTGRES_WEBHOOK, job_id, {"sql": sql, "params": []})
    if err:
        session["retailer"] = None
        session["flags"]["retailer_identification"] = f"webhook_error: {err}"
        await stage_date_config(session_id)
    else:
        session["status"] = "awaiting_postgres"
        session["_pending_jobs"].add(job_id)


async def stage_write_audit(session_id: str):
    """Stage 6b — write discovery result back to file_audit and retailer_configs."""
    session = _sessions[session_id]
    session["stage"] = "writing"
    session["status"] = "running"

    file_audit_id = session.get("file_audit_id")
    retailer      = session.get("retailer")
    result        = session.get("result", {})

    if not file_audit_id:
        # No audit ID — skip DB writes, go straight to assemble
        await stage_assemble(session_id)
        return

    # Determine status for file_audit
    if retailer:
        audit_status = "discovery_complete"
    else:
        audit_status = "pending_review"

    sql = build_update_file_audit_sql(file_audit_id, result, retailer, audit_status)

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {
        "session_id": session_id,
        "stage":      "write_audit",
        "created_at": time.time(),
    }

    err = fire_webhook(N8N_POSTGRES_WEBHOOK, job_id, {"sql": sql, "params": []})
    if err:
        session["errors"].append(f"Failed to write to file_audit: {err}")
        await stage_assemble(session_id)
    else:
        session["status"] = "awaiting_postgres"
        session["_pending_jobs"].add(job_id)

async def stage_date_config(session_id: str):
    """Stage 4 — date config. Pure Python, AI only if year ambiguous."""
    session = _sessions[session_id]
    session["stage"] = "dating"
    session["status"] = "running"
    pending = 0

    for sheet_name in session["qualified_sheets"]:
        grid      = session["grid"].get(sheet_name, {})
        date_axis = grid.get("date_axis", {})
        anchors   = grid.get("year_anchors", [])

        if not date_axis:
            session["date_config"][sheet_name] = {"error": "No date axis"}
            continue

        # Year is present — no AI needed
        if date_axis.get("year_present"):
            year_value = None
            for a in anchors:
                y = str(a.get("value", ""))[:4]
                if y.isdigit():
                    year_value = int(y)
                    break
            session["date_config"][sheet_name] = {
                "date_format":            date_axis["format"],
                "year_present":           True,
                "year_value":             year_value,
                "year_inference_strategy": "embedded_in_dates",
                "year_boundary_detected": date_axis.get("year_boundary_detected", False),
                "normalize_to":           "week_ending_saturday",
                "source":                 "python",
            }
            continue

        # Collect year anchors from other qualified sheets for cross-sheet context
        cross_sheet_anchors = []
        for other_sheet in session["qualified_sheets"]:
            if other_sheet != sheet_name:
                other_anchors = session["grid"].get(other_sheet, {}).get("year_anchors", [])
                cross_sheet_anchors.extend(other_anchors)

        # Year missing or boundary — send to AI
        prompt = build_date_prompt(date_axis, anchors, sheet_name,
                                   filename=session["filename"],
                                   cross_sheet_anchors=cross_sheet_anchors)
        job_id = str(uuid.uuid4())
        _jobs[job_id] = {
            "session_id": session_id,
            "stage":      "date_config",
            "sheet_name": sheet_name,
            "date_axis":  date_axis,
        }
        err = fire_webhook(N8N_AI_WEBHOOK, job_id, {"prompt": prompt})
        if err:
            session["date_config"][sheet_name] = {"error": f"Webhook error: {err}"}
        else:
            session["status"] = "awaiting_ai"
            session["_pending_jobs"].add(job_id)
            pending += 1

    if pending == 0:
        await stage_multisheet(session_id)


async def stage_multisheet(session_id: str):
    """Stage 5 — multi-sheet flag. Pure Python."""
    session   = _sessions[session_id]
    qualified = session["qualified_sheets"]

    multiple = len(qualified) > 1
    session["flags"]["multiple_sales_sheets_detected"] = multiple
    if multiple:
        session["flags"]["multiple_sheets_note"] = (
            f"{len(qualified)} qualified sheets detected. "
            "Next stage should compare actual data to determine combine vs dedup."
        )

    await stage_write_audit(session_id)


async def stage_assemble(session_id: str):
    """Stage 6 — assemble final config."""
    session = _sessions[session_id]
    session["stage"] = "complete"
    session["status"] = "complete"

    session["result"] = {
        "filename":        session["filename"],
        "file_hash":       session["file_hash"],
        "qualified_sheets": session["qualified_sheets"],
        "qualify_results": session["qualify_results"],
        "grid":            session["grid"],
        "column_mapping":  session["column_mapping"],
        "date_config":     session["date_config"],
        "flags":           session["flags"],
    }


# ─────────────────────────────────────────────
# RESPONSE ENDPOINT — shared callback
# ─────────────────────────────────────────────

@app.post("/response/{job_id}")
async def webhook_response(job_id: str, request_body: dict, background_tasks: BackgroundTasks):
    if job_id not in _jobs:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found or already completed")

    job        = _jobs.pop(job_id)
    session_id = job["session_id"]
    stage      = job["stage"]

    if session_id not in _sessions:
        return JSONResponse(content={"status": "ok", "note": "session already closed"})

    session = _sessions[session_id]
    session["_pending_jobs"].discard(job_id)
    # Only set running if there are still pending jobs — otherwise stage handler sets the next status
    if session["_pending_jobs"]:
        session["status"] = "awaiting_ai"

    # Store result by stage
    if stage == "qualify":
        sheet_name = job["sheet_name"]
        raw = request_body
        # Parse AI response
        text = raw.get("text", "") or raw.get("response", "") or ""
        if isinstance(text, str):
            clean = text.replace("```json", "").replace("```", "").strip()
            try:
                verdict = json.loads(clean)
            except:
                verdict = {"disqualified": None, "reason": f"Parse error: {text[:100]}"}
        else:
            verdict = raw
        verdict["source"] = "ai"
        session["qualify_results"][sheet_name] = verdict

        if not session["_pending_jobs"]:
            background_tasks.add_task(advance_from_qualify, session_id)

    elif stage == "postgres_sku":
        session["postgres_results"] = request_body
        if not session["_pending_jobs"]:
            background_tasks.add_task(advance_from_postgres, session_id)

    elif stage == "classify_sheet":
        sheet_name = job["sheet_name"]
        raw = request_body
        text = raw.get("text", "") or ""
        if isinstance(text, str):
            clean = text.replace("```json", "").replace("```", "").strip()
            try:
                result = json.loads(clean)
            except:
                result = {"error": f"Parse error: {text[:100]}"}
        else:
            result = raw
        result["source"] = "ai"
        session["column_mapping"][sheet_name] = result.get("columns", result)

        if not session["_pending_jobs"]:
            background_tasks.add_task(stage_date_config, session_id)

    elif stage == "date_config":
        sheet_name = job["sheet_name"]
        raw = request_body
        text = raw.get("text", "") or ""
        if isinstance(text, str):
            clean = text.replace("```json", "").replace("```", "").strip()
            try:
                result = json.loads(clean)
            except:
                result = {"error": f"Parse error: {text[:100]}"}
        else:
            result = raw
        result["normalize_to"] = "week_ending_saturday"
        result["source"]       = "ai"
        session["date_config"][sheet_name] = result

        if not session["_pending_jobs"]:
            background_tasks.add_task(stage_multisheet, session_id)

    elif stage == "identify_retailer":
        # Postgres returned retailer match results
        rows = request_body.get("matches", [])
        if rows and int(rows[0].get("matches", 0)) >= 3:
            session["retailer"] = rows[0].get("retailer")
            session["flags"]["retailer_identification"] = f"confirmed: {session['retailer']} ({rows[0].get('matches')} matches)"
        else:
            session["retailer"] = None
            session["flags"]["retailer_identification"] = "unconfirmed — fewer than 3 matches"

        if not session["_pending_jobs"]:
            background_tasks.add_task(stage_date_config, session_id)

    elif stage == "write_audit":
        # Postgres confirmed the file_audit update
        # If retailer identified and no active config, insert into retailer_configs
        retailer      = session.get("retailer")
        file_audit_id = session.get("file_audit_id")
        result        = session.get("result", {})

        if retailer and file_audit_id:
            sql = build_insert_retailer_config_sql(retailer, file_audit_id, result)
            job_id2 = str(uuid.uuid4())
            _jobs[job_id2] = {
                "session_id": session_id,
                "stage":      "write_config",
                "created_at": time.time(),
            }
            err = fire_webhook(N8N_POSTGRES_WEBHOOK, job_id2, {"sql": sql, "params": []})
            if err:
                session["errors"].append(f"Failed to write retailer_config: {err}")
                background_tasks.add_task(stage_assemble, session_id)
            else:
                session["status"] = "awaiting_postgres"
                session["_pending_jobs"].add(job_id2)
        else:
            background_tasks.add_task(stage_assemble, session_id)

    elif stage == "write_config":
        # Config row inserted — done
        if not session["_pending_jobs"]:
            background_tasks.add_task(stage_assemble, session_id)

    return JSONResponse(content={"status": "ok", "job_id": job_id})


# ─────────────────────────────────────────────
# ANALYZE ENDPOINT
# ─────────────────────────────────────────────

def file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


@app.post("/analyze")
async def analyze(
    files: List[UploadFile] = File(...),
    file_audit_id: str = None,
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    if not files:
        raise HTTPException(status_code=400, detail="At least one file required")

    upload = files[0]

    if not upload.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail=f"'{upload.filename}' is not an Excel file")

    data = await upload.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot open {upload.filename}: {e}")

    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets[sheet_name] = list(ws.iter_rows(values_only=True))

    session_id = str(uuid.uuid4())
    _sessions[session_id] = {
        "stage":            "accepted",
        "status":           "running",
        "filename":         upload.filename,
        "file_hash":        file_hash(data),
        "file_audit_id":    file_audit_id,
        "retailer":         None,
        "qualified_sheets": [],
        "qualify_results":  {},
        "grid":             {},
        "postgres_results": {},
        "column_mapping":   {},
        "date_config":      {},
        "flags":            {},
        "errors":           [],
        "result":           None,
        "_sheets":          sheets,
        "_pending_jobs":    set(),
        "created_at":       time.time(),
    }

    background_tasks.add_task(stage_qualify, session_id)

    return JSONResponse(content={
        "status":     "accepted",
        "session_id": session_id,
        "sheets":     list(sheets.keys()),
        "poll_url":   f"/analyze/status/{session_id}",
    })


@app.get("/analyze/status/{session_id}")
async def analyze_status(session_id: str):
    if session_id not in _sessions:
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")

    session = _sessions[session_id]

    # Don't expose internal state
    return JSONResponse(content={
        "session_id": session_id,
        "stage":      session.get("stage", "unknown"),
        "status":     session.get("status", "unknown"),
        "filename":   session["filename"],
        "result":     session["result"],
    })


@app.get("/health")
def health():
    return {"status": "ok", "version": "3.0.0"}

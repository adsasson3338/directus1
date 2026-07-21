# -*- coding: utf-8 -*-
"""
discovery.py - Discovery pipeline: qualify, locate, identify, date, write audit.
"""
from fastapi import APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse
import asyncio
import openpyxl
import base64
import hashlib
import io
import re
import uuid
import json
import time
from datetime import datetime, date, timedelta

from shared import (
    _sessions, _jobs,
    call_postgres, call_ai, fire_fetch_file_webhook,
    normalize_to_saturday,
    _validate_uuid, _sql_escape,
    build_fetch_audit_row_sql,
    load_date_patterns, match_known_patterns, normalize_header_shape,
    compute_date_from_match, build_insert_date_pattern_sql,
    build_fetch_existing_patterns_for_dedup_sql,
    build_date_map, resolve_date_header,
    build_sweep_fail_row_sql,
)

# ---------------------------------------------
# AI RESPONSE PARSER
# ---------------------------------------------

def parse_ai_response(text: str) -> str:
    """Strip thinking blocks and code fences from AI response before JSON parsing."""
    if not isinstance(text, str):
        return text
    # Strip <think>...</think> blocks (Qwen and other thinking models)
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL)
    # Strip ```json and ``` fences
    text = text.replace("```json", "").replace("```", "")
    return text.strip()


router = APIRouter()

# ---------------------------------------------
# DATE FORMAT PATTERN LIBRARY
# ---------------------------------------------
# The pattern cache, matching, and computation functions now live in
# shared.py (load_date_patterns, match_known_patterns, normalize_header_shape,
# compute_date_from_match, build_insert_date_pattern_sql) - imported above -
# because ingestion.py needs the exact same logic to compute week_ending
# dates for real sales rows, not just to decide which columns are date
# columns. One canonical implementation, used by both pipelines.

YEAR_RE = re.compile(r"\b(20\d{2})\b")


def classify_cell(val) -> str:
    if val is None:               return "empty"
    if isinstance(val, bool):     return "bool"
    if isinstance(val, datetime): return "datetime"
    if isinstance(val, int):      return "integer"
    if isinstance(val, float):
        return "float_0_to_1" if 0 <= val <= 1 else "float_above_1"
    if isinstance(val, str):
        match = match_known_patterns(val)
        if match:
            return match.get("format_description") or "known_date_pattern"
        return "string"
    return "unknown"

def is_date_like(val) -> bool:
    if isinstance(val, datetime):
        return True
    return match_known_patterns(val) is not None if isinstance(val, str) else False



MONTH_MAP = {
    'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
    'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12
}

def extract_month_sequence(row, col_list: list) -> list:
    """
    Extract the tracking month for each date column in order.
    For date-range strings, uses the END month (the resolved date).
    For fiscal week labels, uses the month number.
    For datetime objects, uses the month directly.
    Returns a list of ints (1-12) or None for unparseable values.
    """
    months = []
    for ci in col_list:
        if ci >= len(row) or row[ci] is None:
            months.append(None)
            continue
        val = row[ci]
        if isinstance(val, datetime):
            months.append(val.month)
            continue
        s = str(val).strip()
        # date_range: use END month
        m = re.match(r'^(\d{1,2})/(\d{1,2})[-\u2013](\d{1,2})/(\d{1,2})$', s)
        if m:
            months.append(int(m.group(3)))
            continue
        # fiscal week label
        mf = re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+wk',
                      s, re.IGNORECASE)
        if mf:
            months.append(MONTH_MAP[mf.group(1).lower()[:3]])
            continue
        months.append(None)
    return months


def compute_year_start(month_sequence: list, year_value: int) -> int:
    """
    Determine the starting year for the date map from the month sequence.

    52-column files (full year): majority rule - find the starting year
    such that the year with the most columns equals year_value.

    <52-column files (partial year):
    - Q4 months before January -> prior-year tail -> year_start = year_value - 1
    - Q1 month at start -> current year -> year_start = year_value
    - H2 opening month fallback -> year_start = year_value - 1
    """
    valid = [m for m in month_sequence if m is not None]
    if not valid:
        return year_value

    if len(valid) == 52:
        for try_start in [year_value - 1, year_value]:
            counts = {}
            current = try_start
            prev = None
            for m in valid:
                if prev is not None and m < prev:
                    current += 1
                counts[current] = counts.get(current, 0) + 1
                prev = m
            if counts and max(counts, key=counts.get) == year_value:
                return try_start
        return year_value

    # <52 columns
    first_month   = valid[0]
    jan_positions = [i for i, m in enumerate(valid) if m == 1]
    first_jan     = jan_positions[0] if jan_positions else None
    q4_before_jan = first_jan is not None and any(
        m in (10, 11, 12) for m in valid[:first_jan]
    )
    if q4_before_jan:
        return year_value - 1
    if first_month in (1, 2, 3):
        return year_value
    return year_value - 1 if first_month > 6 else year_value


EMBEDDED_PATTERNS = [
    ("integer_dash_description",        re.compile(r'^(\d{4,10})\s*[--]\s*(.{3,})$')),
    ("alphanumeric_space_description",  re.compile(r'^([A-Z0-9]{2,}-[A-Z0-9\-]{2,})\s+(.{3,})$', re.IGNORECASE)),
    ("description_space_alphanumeric",  re.compile(r'^(.{3,})\s{2,}([A-Z0-9]{2,}-[A-Z0-9\-]{2,})\s*$', re.IGNORECASE)),
    ("description_space_nodash_sku",    re.compile(r'^(.{5,})\s+([A-Z]{2,}\d{3,})\s*$', re.IGNORECASE)),
]

# ---------------------------------------------
# GRID DETECTION
# ---------------------------------------------

# Used in two places: excluding inventory-labeled columns from the sales
# date axis (find_date_axis, below), and gating column_mapping's
# "inventory" classification (stage_schema_classify) so AI's judgment on
# this specific, high-stakes label isn't trusted on its own - a column
# only counts as inventory if its header actually contains one of these
# words, regardless of what AI's free-form reasoning concluded. Real
# production example: the same "Current Week" column, same file, got
# classified "other" on one discovery run and "inventory" on another -
# AI reasoning about intent is not reliable enough for a classification
# that determines which TABLE the data gets written into.
INVENTORY_LABELS = {
    'inv', 'inventory', 'on order', 'onorder', 'dc inv', 'store inv',
    'total inv', 'on hand', 'onhand', 'stock', 'qty on hand',
    'remaining', 'order qty', 'open order'
}


def find_date_axis(rows) -> dict | None:
    best = {"row": None, "count": 0, "cols": [], "samples": [], "format": None}
    for row_idx, row in enumerate(rows[:15]):
        date_cols = [(ci, v) for ci, v in enumerate(row) if is_date_like(v)]
        if len(date_cols) > best["count"]:
            formats = set(classify_cell(v) for _, v in date_cols)
            best = {
                "row": row_idx, "count": len(date_cols),
                "cols": [ci for ci, _ in date_cols],
                "samples": [str(v) for _, v in date_cols[:8]],
                "format": list(formats)[0] if len(formats) == 1 else "mixed",
            }
    if best["count"] < 2:
        return None

    cols = best["cols"]
    year_present = best["format"] == "datetime" or any(YEAR_RE.search(s) for s in best["samples"])

    interleaved = False
    if len(cols) >= 3:
        gaps = [cols[i+1] - cols[i] for i in range(len(cols)-1)]
        interleaved = len(set(gaps)) == 1 and gaps[0] == 2

    # Year boundary detection: reuse the canonical month-extraction function
    # rather than re-deriving months here with a second, independent regex.
    # Works for any pattern extract_month_sequence understands (currently
    # date_range and fiscal_week); silently contributes nothing for formats
    # it doesn't yet handle - year_boundary_detected just won't fire for a
    # brand-new pattern until someone extends extract_month_sequence for it,
    # which is a bounded, honest limitation rather than a silent wrong answer.
    candidate_months = extract_month_sequence(rows[best["row"]], cols)
    year_boundary_detected = any(
        a == 12 and b == 1
        for a, b in zip(candidate_months, candidate_months[1:])
        if a is not None and b is not None
    )

    # Filter out inventory/summary columns that have date headers but are not sales
    # Check the row above the date axis for inventory-related labels
    # (INVENTORY_LABELS is now a module-level constant, shared with the
    # column_mapping gate in stage_schema_classify)
    if best["row"] > 0:
        label_row = rows[best["row"] - 1]
        cols = [
            ci for ci in cols
            if not any(
                inv in (str(label_row[ci]).strip().lower() if ci < len(label_row) and label_row[ci] else "")
                for inv in INVENTORY_LABELS
            )
        ]
        # Rebuild samples from the filtered cols directly, rather than
        # zipping against the pre-truncated (display-only) samples list -
        # zip() silently stops at the shorter iterable, which previously
        # capped `cols` itself at 8 regardless of how many real date
        # columns existed (a pre-existing bug, not introduced by this pass).
        row = rows[best["row"]]
        best["samples"] = [str(row[ci]) for ci in cols[:8] if ci < len(row)]

    # Extract ordered tracking months for year_boundary_detected detection (fiscal weeks)
    # and for stage_date_config (via schema path - not from this function's return value)
    month_sequence = extract_month_sequence(rows[best["row"]], cols) if year_boundary_detected else []

    return {
        "row": best["row"], "col_count": len(cols), "cols": cols,
        "sample_values": best["samples"], "format": best["format"],
        "year_present": year_present, "interleaved_empty_cols": interleaved,
        "year_boundary_detected": year_boundary_detected,
    }


def find_data_start_row(rows, date_axis_row: int, date_cols: list) -> int:
    for row_idx in range(date_axis_row + 1, min(date_axis_row + 10, len(rows))):
        row = rows[row_idx]
        vals = [row[dc] for dc in date_cols if dc < len(row) and row[dc] is not None]
        numeric = [v for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if numeric:
            return row_idx
    return date_axis_row + 1


def detect_embedded_sku(rows, date_cols: list, data_start_row: int) -> list:
    if not date_cols:
        return []
    left_boundary = min(date_cols)
    embedded_candidates = []

    for col_idx in range(left_boundary):
        col_vals = [
            str(rows[r][col_idx]).strip()
            for r in range(data_start_row, min(data_start_row + 30, len(rows)))
            if col_idx < len(rows[r]) and rows[r][col_idx] is not None
            and isinstance(rows[r][col_idx], str)
        ]
        if len(col_vals) < 3:
            continue

        col_matches = []
        seen = set()
        for pattern_name, pattern_re in EMBEDDED_PATTERNS:
            for val in col_vals:
                if val not in seen:
                    m = pattern_re.match(val)
                    if m:
                        seen.add(val)
                        # description_space_* patterns have description in group(1), SKU in group(2)
                        if pattern_name.startswith("description_space"):
                            sku, desc = m.group(2).strip(), m.group(1).strip()
                        else:
                            sku, desc = m.group(1).strip(), m.group(2).strip()
                        col_matches.append({
                            "raw": val, "sku": sku,
                            "description": desc, "pattern": pattern_name,
                        })
        if col_matches:
            embedded_candidates.append({
                "col": col_idx, "total_values": len(col_vals),
                "matched_values": len(col_matches), "extractions": col_matches,
            })
    return embedded_candidates



# ---------------------------------------------
# QUALIFY SIGNALS
# ---------------------------------------------

def extract_qualify_signals(rows: list, sheet_name: str, filename: str) -> dict:
    integer_count = float_above_count = float_0_to_1_count = 0
    crosshair_sample = []

    date_axis = find_date_axis(rows)
    if date_axis:
        date_cols  = date_axis["cols"][:8]
        data_start = find_data_start_row(rows, date_axis["row"], date_axis["cols"])
        for row in rows[data_start:data_start + 50]:
            for dc in date_cols:
                if dc >= len(row) or row[dc] is None:
                    continue
                v = row[dc]
                if isinstance(v, bool):
                    continue
                if isinstance(v, int):
                    integer_count += 1
                    if len(crosshair_sample) < 40:
                        crosshair_sample.append(v)
                elif isinstance(v, float) and v > 1:
                    float_above_count += 1
                    if len(crosshair_sample) < 40:
                        crosshair_sample.append(v)
                elif isinstance(v, float) and 0 < v <= 1:
                    float_0_to_1_count += 1
                    if len(crosshair_sample) < 40:
                        crosshair_sample.append(v)
        total = integer_count + float_above_count + float_0_to_1_count
        if total == 0:
            dominant_type = None
        elif float_above_count / total > 0.7:
            dominant_type = "float_above_1"
        elif float_0_to_1_count / total > 0.7:
            dominant_type = "float_0_to_1"
        elif integer_count / total > 0.7:
            dominant_type = "integer"
        else:
            dominant_type = "mixed"
    else:
        for row in rows[:50]:
            for v in row:
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    if len(crosshair_sample) < 40:
                        crosshair_sample.append(v)
            if len(crosshair_sample) >= 40:
                break
        dominant_type = None

    column_labels = set()
    for row in rows[:6]:
        for val in row:
            if isinstance(val, str) and val.strip() and len(val.strip()) < 60:
                column_labels.add(val.strip())

    inline_strings = set()
    for row in rows[6:20]:
        for val in row:
            if isinstance(val, str) and val.strip() and len(val.strip()) < 80:
                inline_strings.add(val.strip())

    return {
        "sheet_name":       sheet_name,
        "filename":         filename,
        "dominant_type":    dominant_type,
        "crosshair_sample": crosshair_sample,
        "column_labels":    sorted(column_labels, key=len)[:10],
        "inline_strings":   sorted(inline_strings, key=len)[:10],
        "date_col_count":   len(date_axis["cols"]) if date_axis else 0,
    }


def build_qualify_prompt(signals: dict) -> str:
    return f"""You are the gatekeeper of a retail sales data ingestion pipeline. A sheet has arrived and you must determine if it should be disqualified.

A sheet should be disqualified if it does NOT contain unit sales data. Unit sales data has integer values at the intersection of product identifiers and date columns.

Disqualify if any of the following are true:
- Crosshair values are decimals above 1 - dollar revenue
- Crosshair values are between 0 and 1 - percentage metrics
- Column labels contain $$$ or $$ - dollar sheet
- Sheet name contains CFP, Forecast, FCST, Projection, Order
- Vocabulary contains forecast or projection language

EVIDENCE:
Sheet name: {signals["sheet_name"]}
Filename: {signals["filename"]}
Detected date columns: {signals.get("date_col_count", 0)}
Crosshair sample values: {signals["crosshair_sample"]}
Dominant crosshair type: {signals["dominant_type"]}
Column labels: {signals["column_labels"]}
Inline strings: {signals["inline_strings"]}

Respond with JSON only:
{{"disqualified": true or false, "reason": "one sentence explanation"}}"""


# ---------------------------------------------
# POSTGRES SQL BUILDER
# ---------------------------------------------

def build_sku_lookup_sql(candidates: list) -> tuple[str, list]:
    """
    Build case-insensitive exact match query against inventory_view.
    Values embedded inline - n8n Postgres node does not support $1 params.
    Returns (sql, []) - empty params list.
    """
    # Safely quote each candidate for inline SQL
    quoted = ", ".join(
        "'" + str(c).replace("'", "''") + "'"
        for c in candidates
        if c and str(c).strip()
    )
    if not quoted:
        quoted = "''"

    sql = f"""SELECT inventory_sku, base_model, base_variant, description, upc
FROM inventory_view
WHERE UPPER(inventory_sku) = ANY(ARRAY[{quoted}]::text[])
   OR UPPER(base_variant)  = ANY(ARRAY[{quoted}]::text[])
   OR UPPER(base_model)    = ANY(ARRAY[{quoted}]::text[])"""

    return sql, []


# ---------------------------------------------
# COLUMN CLASSIFY PROMPT
# ---------------------------------------------


def build_date_prompt(date_axis: dict, year_anchors: list, sheet_name: str,
                       filename: str = "", cross_sheet_anchors: list = None) -> str:
    cross_context = ""
    if cross_sheet_anchors:
        cross_context = f"\nYear anchors from other sheets in same file: {cross_sheet_anchors}"

    return f"""You are configuring the date settings for a retail sales sheet.

Determine the document year and week convention from the evidence below.
The file was received as: {filename}

Sheet: {sheet_name}
Date axis format: {date_axis["format"]}
Sample date values: {date_axis["sample_values"]}
Year anchors in this sheet: {year_anchors}{cross_context}
Detected date columns: {date_axis.get("date_col_count", 0)}

Respond with JSON only:
{{"year_value": 2026, "week_convention": "what convention the source uses", "year_inference_strategy": "one sentence"}}"""



# ---------------------------------------------
# RETAILER IDENTIFICATION SQL
# ---------------------------------------------



def build_dedup_check_sql(file_hash: str) -> str:
    """
    Check if this file has already been seen - genuinely successful or
    still-in-progress statuses count as a duplicate; 'failed' does not.

    A file that previously failed is exactly the case where a retry
    should be ALLOWED, not blocked - treating a failed prior attempt as
    "already handled" (the old behavior: any status at all counted as a
    match) meant a file that got stuck and failed once could never be
    resubmitted at all, even after a human reset its own row back to
    'received' - a brand new submission would still find the OLD failed
    row under the same hash and skip before even trying.
    """
    return f"""
SELECT id, status, filename
FROM file_audit
WHERE file_hash = '{file_hash}'
  AND status != 'failed'
LIMIT 1
""".strip()

def build_retailer_identify_sql(retailer_sku_candidates: list) -> str:
    """
    Query retailer_sku_map to identify retailer from SKU candidates.
    Returns retailer and match count - 3+ matches confirms identity.
    """
    quoted = ", ".join(
        "'" + str(c).replace("'", "''") + "'"
        for c in retailer_sku_candidates
        if c and str(c).strip()
    )
    if not quoted:
        quoted = "''"

    return f"""
SELECT retailer, COUNT(*) as matches
FROM retailer_sku_map
WHERE UPPER(retailer_sku) = ANY(ARRAY[{quoted}]::text[])
  AND active = true
GROUP BY retailer
ORDER BY matches DESC
LIMIT 1
""".strip()


def sql_escape(v) -> str:
    """Escape a value for safe inline SQL embedding."""
    return str(v).replace("'", "''")


def json_safe(v) -> str:
    """Serialize to JSON with unicode preserved, then SQL-escape."""
    return sql_escape(json.dumps(v, ensure_ascii=False))


def build_insert_retailer_config_sql(retailer: str) -> str:
    """
    Insert a pending_review config row for a newly identified retailer.
    Only fires once - skipped if retailer already exists.
    """
    safe_retailer = _sql_escape(retailer) if hasattr(retailer, "__class__") else str(retailer).replace("'", "''")
    return f"""
INSERT INTO retailer_configs (retailer, status, file_set_size)
SELECT '{safe_retailer}', 'pending_review', 1
WHERE NOT EXISTS (
    SELECT 1 FROM retailer_configs
    WHERE UPPER(retailer) = UPPER('{safe_retailer}')
)
""".strip()


def build_retailer_identify_prompt(filename: str, sheets: list, header_strings: list) -> str:
    """
    Build a prompt to identify retailer from file/sheet clues when Postgres has no match.
    """
    return f"""You are identifying which retailer sent a sales file to a supplier.

Known retailers: Walgreens, Target, Staples, Walmart, CVS, Rite Aid, Best Buy, Amazon, Costco, BJ's, Sam's Club.

CLUES:
Filename: {filename}
Sheet names: {sheets}
Header strings found in sheets: {header_strings[:20]}

Only identify the retailer if the clues EXPLICITLY mention the retailer name or a well-known retailer-specific identifier (e.g. "Walgreens", "WIC#", "DPCI", "Staples").

DO NOT GUESS. If the retailer name is not explicitly present in the filename, sheet names, or headers, return null.
A generic filename like "D56_FCSTs_2026.xlsx" with no retailer name is NOT sufficient - return null.

Respond with JSON only:
{{"retailer": "retailer name or null", "confidence": "high/medium/low", "reason": "one sentence"}}"""


def build_query_retailer_config_sql(retailer: str) -> str:
    """Query active retailer config to get file_set_size."""
    retailer_safe = retailer.replace("'", "''")
    return f"""
SELECT file_set_size, id
FROM retailer_configs
WHERE UPPER(retailer) = UPPER('{retailer_safe}')
  AND status = 'active'
LIMIT 1
""".strip()


def build_file_set_key(retailer: str, date_config: dict, grid: dict = None) -> str:
    """
    Build a file set key from retailer and the latest week-ending date in the file.
    Uses date_config (which has year_start, year_value, date_cols) to resolve dates
    correctly — the same logic used by ingestion.
    Format: RETAILER_YYYY-MM-DD
    """
    from datetime import date as _date, timedelta

    retailer_clean = re.sub(r"[^A-Z0-9]", "_", retailer.upper()).strip("_")

    latest_date = None

    for sheet_cfg in date_config.values():
        year_start = sheet_cfg.get("year_start") or sheet_cfg.get("year_value")
        if not year_start:
            continue

        # Use month_sequence from grid if available to resolve dates correctly
        sheet_name = None
        for sn, sc in date_config.items():
            if sc is sheet_cfg:
                sheet_name = sn
                break

        month_seq = []
        last_sample_val = None
        if grid and sheet_name:
            date_axis_g          = grid.get(sheet_name, {}).get("date_axis", {})
            month_seq            = date_axis_g.get("month_sequence", [])
            last_active_sample   = date_axis_g.get("last_active_sample_val")
            last_sample_val      = last_active_sample or date_axis_g.get("last_sample_val")

        # First try: parse last_sample_val directly (most accurate)
        if last_sample_val:
            s = str(last_sample_val).strip()
            # MM/DD/YY or MM/DD/YYYY
            m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
            if m:
                mo, dy, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
                yr = 2000 + yr if yr < 100 else yr
                try:
                    parsed = _date(yr, mo, dy)
                    parsed = parsed + timedelta(days=(5 - parsed.weekday()) % 7)
                    if latest_date is None or parsed > latest_date:
                        latest_date = parsed
                except ValueError:
                    pass
            # MM/DD-MM/DD range — use end date with year_start + monotonic rule
            m2 = re.match(r"^(\d{1,2})/(\d{1,2})[--](\d{1,2})/(\d{1,2})$", s)
            if m2 and not latest_date:
                end_mo, end_dy = int(m2.group(3)), int(m2.group(4))
                eff_year = year_start
                if month_seq:
                    current = year_start
                    prev_m = None
                    for mo in month_seq:
                        if mo is None: continue
                        if prev_m is not None and mo < prev_m:
                            current += 1
                        prev_m = mo
                    eff_year = current
                try:
                    parsed = _date(eff_year, end_mo, end_dy)
                    if latest_date is None or parsed > latest_date:
                        latest_date = parsed
                except ValueError:
                    pass
            # fiscal_week_label — use month_sequence last month
            mf = re.match(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+wk\s*(\d+)$", s, re.IGNORECASE)
            if mf and not latest_date:
                last_month_num = MONTH_MAP[mf.group(1).lower()[:3]]
                wk_num = int(mf.group(2))
                eff_year = year_start
                if month_seq:
                    current = year_start
                    prev_m = None
                    for mo in month_seq:
                        if mo is None: continue
                        if prev_m is not None and mo < prev_m:
                            current += 1
                        prev_m = mo
                    eff_year = current
                try:
                    first = _date(eff_year, last_month_num, 1)
                    approx = first + timedelta(days=(wk_num - 1) * 7)
                    parsed = approx + timedelta(days=(5 - approx.weekday()) % 7)
                    if latest_date is None or parsed > latest_date:
                        latest_date = parsed
                except ValueError:
                    pass

        elif month_seq:
            # Walk month_sequence with monotonic year rule to find the last date
            current_year = year_start
            prev_m = None
            last_month = None
            last_year  = None
            for m in month_seq:
                if m is None:
                    continue
                if prev_m is not None and m < prev_m:
                    current_year += 1
                last_month = m
                last_year  = current_year
                prev_m = m

            if last_month and last_year:
                try:
                    # Approximate last week of last month as the 4th Saturday
                    first = _date(last_year, last_month, 1)
                    approx = first + timedelta(days=3 * 7)
                    parsed = approx + timedelta(days=(5 - approx.weekday()) % 7)
                    if latest_date is None or parsed > latest_date:
                        latest_date = parsed
                except ValueError:
                    pass
        else:
            # Fall back to sample_values from grid if no month_sequence
            if not grid or not sheet_name:
                continue
            samples = grid.get(sheet_name, {}).get("date_axis", {}).get("sample_values", [])
            year_boundary = sheet_cfg.get("year_boundary_detected", False)
            year_value    = sheet_cfg.get("year_value", year_start)
            for sample in samples:
                s = str(sample).strip()
                m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
                if m:
                    mo, dy, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
                    yr = 2000 + yr if yr < 100 else yr
                    try:
                        parsed = _date(yr, mo, dy)
                        if latest_date is None or parsed > latest_date:
                            latest_date = parsed
                    except ValueError:
                        pass

    if latest_date:
        return f"{retailer_clean}_{latest_date.isoformat()}"

    # Fallback
    year_value = next((v.get("year_value") for v in date_config.values() if v.get("year_value")), _date.today().year)
    week_num = _date.today().isocalendar()[1]
    return f"{retailer_clean}_{year_value}-W{week_num:02d}"


def build_claim_file_audit_sql(file_audit_id: str) -> str:
    """
    Atomically flip status received -> analyzing, conditioned on the row
    still being 'received'. Returns the updated id if this call won the
    race, or zero rows if another concurrent poll already claimed it.
    Prevents the same file_audit row from being picked up twice by
    overlapping polling runs while discovery is still in progress.
    """
    return f"""
UPDATE file_audit
SET status = 'analyzing', updated_at = now()
WHERE id = '{file_audit_id}'
  AND status = 'received'
RETURNING id
""".strip()


def build_update_file_audit_full_sql(file_audit_id: str, discovery_result: dict,
                                      retailer: str | None, status: str,
                                      file_set_key: str | None,
                                      file_hash: str | None = None,
                                      filename: str | None = None,
                                      resolved_dates: dict | None = None) -> str:
    """Update file_audit - writes discovery_result plus dedicated columns for easy querying."""
    result_b64 = base64.b64encode(json.dumps(discovery_result, ensure_ascii=False).encode()).decode()

    # Extract the dedicated columns directly
    qs_b64 = base64.b64encode(json.dumps(discovery_result.get("qualified_sheets", []), ensure_ascii=False).encode()).decode()
    cm_b64 = base64.b64encode(json.dumps(discovery_result.get("column_mapping",   {}), ensure_ascii=False).encode()).decode()
    dc_b64 = base64.b64encode(json.dumps(discovery_result.get("date_config",      {}), ensure_ascii=False).encode()).decode()
    # resolved_dates lives ONLY here, a plain top-level column - never
    # duplicated inside discovery_result/date_config's nested JSON. Kept
    # flat and minimal on purpose: {"Sheet1": {"5": "2026-01-03", ...}},
    # nothing else - not year_start/week_convention/inference strategy/etc,
    # which stay in date_config where they belong. This is the one field
    # with a simple, checkable invariant (its count should match date_cols'
    # count), which is exactly what made it worth pulling out on its own.
    rd_b64 = base64.b64encode(json.dumps(resolved_dates or {}, ensure_ascii=False).encode()).decode()

    retailer_val = f"'{sql_escape(retailer)}'" if retailer else "NULL"
    key_val      = f"'{sql_escape(file_set_key)}'" if file_set_key else "NULL"
    hash_val     = f"'{file_hash}'" if file_hash else "NULL"
    fname_val    = f"'{sql_escape(filename)}'" if filename else "NULL"

    return f"""
UPDATE file_audit
SET discovery_result = convert_from(decode('{result_b64}', 'base64'), 'UTF8')::jsonb,
    qualified_sheets = convert_from(decode('{qs_b64}', 'base64'), 'UTF8')::jsonb,
    column_mapping   = convert_from(decode('{cm_b64}', 'base64'), 'UTF8')::jsonb,
    date_config      = convert_from(decode('{dc_b64}', 'base64'), 'UTF8')::jsonb,
    resolved_dates   = convert_from(decode('{rd_b64}', 'base64'), 'UTF8')::jsonb,
    retailer         = {retailer_val},
    status           = '{status}',
    file_set_key     = {key_val},
    file_hash        = {hash_val},
    filename         = {fname_val},
    updated_at       = now()
WHERE id = '{file_audit_id}'
""".strip()


def extract_column_schema(ws, sheet_name: str) -> dict:
    """
    Extract raw column schema from worksheet.
    Detects data block start, header rows, merged cells, borders.
    Returns per-column: header stack, borders, data stats.
    """
    max_col = ws.max_column
    max_row  = ws.max_row

    # Build merged cell map
    merge_map = {}
    for merge in ws.merged_cells.ranges:
        val = ws.cell(merge.min_row, merge.min_col).value
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                merge_map[(row, col)] = {
                    "value": val,
                    "range": str(merge),
                    "is_top_left": (row == merge.min_row and col == merge.min_col)
                }

    def get_cell_value(row, col):
        if (row, col) in merge_map:
            return merge_map[(row, col)]["value"]
        return ws.cell(row, col).value

    # Find data block start - first row with >30% numeric values
    data_start_row = None
    for row_idx in range(1, min(30, max_row + 1)):
        values = [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]
        numeric = sum(1 for v in values if isinstance(v, (int, float)) and not isinstance(v, bool))
        non_null = sum(1 for v in values if v is not None)
        if non_null > 0 and numeric / max(non_null, 1) > 0.3:
            data_start_row = row_idx
            break

    if not data_start_row:
        data_start_row = 2

    header_rows = list(range(1, data_start_row))

    columns = []
    for col_idx in range(1, max_col + 1):
        # Collect header stack
        header_stack = []
        for row_idx in header_rows:
            val = get_cell_value(row_idx, col_idx)
            if val is not None and str(val).strip():
                if isinstance(val, datetime):
                    val = val.strftime('%m/%d/%y')
                is_merged   = (row_idx, col_idx) in merge_map
                merge_parent = merge_map.get((row_idx, col_idx), {}).get("is_top_left", False)
                header_stack.append({
                    "row":          row_idx,
                    "value":        str(val).strip(),
                    "merged":       is_merged,
                    "merge_parent": merge_parent,
                })

        # Border info from last header row
        last_hrow = header_rows[-1] if header_rows else 1
        cell = ws.cell(last_hrow, col_idx)
        borders = {
            "left":   cell.border.left.style if cell.border.left else None,
            "right":  cell.border.right.style if cell.border.right else None,
            "top":    cell.border.top.style if cell.border.top else None,
            "bottom": cell.border.bottom.style if cell.border.bottom else None,
        }

        # Data stats
        data_values = [
            ws.cell(r, col_idx).value
            for r in range(data_start_row, min(data_start_row + 100, max_row + 1))
            if ws.cell(r, col_idx).value is not None
        ]
        total    = len(data_values)
        zeros    = sum(1 for v in data_values if v == 0 or v == 0.0)
        non_zero = [v for v in data_values if v and v != 0]
        pct_zero = round(zeros / total, 2) if total > 0 else None
        sample   = [str(v)[:20] for v in non_zero[:5]]

        types = set()
        for v in data_values[:50]:
            if isinstance(v, bool):     types.add("bool")
            elif isinstance(v, int):    types.add("integer")
            elif isinstance(v, float):  types.add("float")
            elif isinstance(v, str):    types.add("string")
            elif isinstance(v, datetime): types.add("datetime")

        columns.append({
            "col":              col_idx,
            "header_stack":     header_stack,
            "borders":          borders,
            "data_types":       sorted(types),
            "sample_non_zero":  sample,
            "pct_zero":         pct_zero,
            "total_data_rows":  total,
        })

    return {
        "sheet_name":     sheet_name,
        "data_start_row": data_start_row,
        "header_rows":    header_rows,
        "columns":        [c for c in columns if c["header_stack"] or c["sample_non_zero"]],
        "merged_ranges":  [str(m) for m in ws.merged_cells.ranges],
    }


def detect_first_sales_col(schema: dict, date_axis_row: int) -> int:
    """
    Detect the first column (1-based) containing a date-like value in the date axis row.
    """
    for col in sorted(schema.get("columns", []), key=lambda c: c["col"]):
        for h in col.get("header_stack", []):
            if h["row"] == date_axis_row:
                val = str(h.get("value", "")).strip()
                if match_known_patterns(val):
                    return col["col"]
    return 1


def build_date_schema_prompt(schema: dict, filename: str) -> str:
    """
    AI identifies only the semantic label context for the sales date block:
    - What label (if any) appears above the sales columns (e.g. "UNITS", "Sales")
    - What labels signal the END of the sales block (e.g. "INV", "TOTAL", "On Order")
    - What row contains those section labels

    Everything else (date_axis_row, first_sales_col, data_start_row, year_present,
    year_boundary, date_format) is detected by Python from the data directly.
    """
    header_rows = schema.get("header_rows", [])

    # Build compact header grid
    header_grid = []
    for row_num in header_rows:
        row_cells = []
        for col in schema.get("columns", []):
            val = ""
            for h in col.get("header_stack", []):
                if h["row"] == row_num:
                    val = h["value"][:15]
                    break
            if val:
                row_cells.append(f"c{col['col']}={repr(val)}")
        if row_cells:
            header_grid.append(f"Row {row_num}: {', '.join(row_cells[:50])}")

    grid_text = "\n".join(header_grid)

    return f"""You are analyzing the header structure of a retail sales spreadsheet.

File: {filename}
Sheet: {schema["sheet_name"]}

HEADER ROWS:
{grid_text}

The sales date columns are a consecutive block of weekly date values (e.g. "01/03/26", "12/21-12/27", "Feb Wk 1").
They may have a section label in the row above (e.g. "Sales", "UNITS") and end when the section label changes (e.g. "INV", "TOTAL", "On Order").

Identify:
1. section_label_row: row number (1-based) containing section group labels ABOVE the date values (null if none)
2. sales_section_label: the exact label above the SALES date columns (null if no label row)
3. stop_labels: exact labels that signal the END of the sales block (empty list if the sheet has no non-sales sections)

Respond with JSON only:
{{"section_label_row": 4, "sales_section_label": "UNITS", "stop_labels": ["INV", "TOTAL", "On Order"]}}"""


def detect_date_axis_row(schema: dict) -> int:
    """Detect the row containing the most date-like values. Returns 1-based row number."""
    row_date_counts = {}
    for col in schema.get("columns", []):
        for h in col.get("header_stack", []):
            row = h["row"]  # already 1-based
            val = str(h.get("value", "")).strip()
            if match_known_patterns(val):
                row_date_counts[row] = row_date_counts.get(row, 0) + 1
    if not row_date_counts:
        return 1
    return max(row_date_counts, key=row_date_counts.get)


def find_sales_cols_from_schema(schema: dict, date_schema: dict) -> list:
    """
    Python enumerates sales date columns.
    AI provides only section label context (sales_section_label, stop_labels).
    Everything else is detected by Python.
    """
    date_axis_row     = detect_date_axis_row(schema)                         # 1-based, Python
    first_sales_col   = detect_first_sales_col(schema, date_axis_row)        # 1-based, Python
    section_label_row = date_schema.get("section_label_row")                 # 1-based or None, AI
    sales_label       = (date_schema.get("sales_section_label") or "").strip().upper()  # AI
    stop_labels       = [s.strip().upper() for s in date_schema.get("stop_labels", [])]  # AI

    # If AI put section_label_row on the same row as date_axis_row, it got confused —
    # there is no separate section label row, so ignore section label context entirely
    if section_label_row == date_axis_row:
        section_label_row = None
        sales_label       = ""

    # Build col -> header values map
    col_headers = {}
    for col in schema.get("columns", []):
        col_num = col["col"]
        headers = {h["row"]: h["value"] for h in col.get("header_stack", [])}  # already 1-based
        col_headers[col_num] = headers

    def is_date_value(val: str) -> bool:
        return match_known_patterns(val) is not None

    def get_section_label(col_num: int) -> str:
        if not section_label_row:
            return ""
        return col_headers.get(col_num, {}).get(section_label_row, "").strip().upper()

    def get_date_value(col_num: int) -> str:
        return col_headers.get(col_num, {}).get(date_axis_row, "").strip()

    sales_cols = []
    all_col_nums = sorted(col_headers.keys())

    in_sales_block = False
    for col_num in all_col_nums:
        if col_num < first_sales_col:
            continue

        date_val    = get_date_value(col_num)
        sec_label   = get_section_label(col_num)

        # Check if section label signals end of sales block
        if sec_label and stop_labels and sec_label in stop_labels:
            break  # hit a stop label - sales block ended

        # Check if this column has a date value in the date axis row
        if date_val and is_date_value(date_val):
            # Verify section label matches (if sales has a label)
            if sales_label and sec_label and sec_label != sales_label:
                break  # section changed
            sales_cols.append(col_num)
            in_sales_block = True
        elif in_sales_block:
            # Was in sales block - if no date value and no section label, skip (empty col)
            if not date_val and not sec_label:
                continue
            # Otherwise the block has ended
            break

    return sales_cols


# ---------------------------------------------
# NEW-FORMAT DISCOVERY (data-shape prefilter + narrow AI escalation)
# ---------------------------------------------
# When find_sales_cols_from_schema() finds nothing (no header matches any
# approved pattern), the old behavior was to disqualify the sheet outright -
# a silent, confident-looking failure indistinguishable from "this sheet
# genuinely has no sales data" (see the 202601-style Walmart header bug).
#
# Instead: first ask Python which columns are even PLAUSIBLE sales-data
# candidates by data shape alone (numeric-only, reasonably populated) -
# this mechanically excludes blank separator columns, flag/marker columns,
# inventory-snapshot datetimes, and percentage columns before any header
# text is interpreted. Only the small residual that survives this AND
# still doesn't match a known pattern gets escalated to AI - one narrow,
# well-scoped judgment call, not "figure out the whole sheet."

def find_sales_shaped_columns(schema: dict) -> list:
    """
    Returns 1-based column numbers whose underlying DATA looks like real
    sales numbers - numeric-only, reasonably populated. Doesn't try to be
    perfectly precise, just eliminates the obvious non-candidates (blanks,
    text, flags, datetimes) before header text is ever interpreted.
    """
    candidates = []
    for col in schema.get("columns", []):
        types = set(col.get("data_types", []))
        total = col.get("total_data_rows", 0)
        if total < 10:                     # too sparse - likely blank/separator
            continue
        if not types or (types - {"integer", "float"}):  # any non-numeric type present
            continue
        candidates.append(col["col"])
    return candidates


def find_data_start_row_from_rows(rows: list) -> int:
    """
    Raw-rows equivalent of extract_column_schema()'s data_start_row
    detection: first row where >30% of non-null values are numeric. 0-based
    (raw row tuples), unlike extract_column_schema()'s 1-based version.
    Needed so the shape-prefilter below doesn't scan header rows - a header
    STRING sitting in the same column as real numeric data underneath it
    would otherwise break the "this column is numeric-only" check entirely.
    """
    for ri in range(min(30, len(rows))):
        row = rows[ri]
        non_null = [v for v in row if v is not None]
        numeric  = [v for v in non_null if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if non_null and len(numeric) / len(non_null) > 0.3:
            return ri
    return 1


def find_sales_shaped_columns_from_rows(rows: list) -> list:
    """
    Raw-rows equivalent of find_sales_shaped_columns(), for use at QUALIFY
    time - before extract_column_schema()'s richer schema object exists
    (that requires reopening the full, non-read-only workbook, which only
    happens after a sheet has already passed qualify). Without this, a
    genuinely new date format gets disqualified at qualify time and never
    reaches the schema-level escalation at all, since qualify runs first
    and stops the pipeline on rejection.

    Only scans from the detected data-start row onward - scanning header
    rows too would mix header STRING values into the "numeric-only" check
    for every column and cause it to find nothing at all.

    0-based column indices (raw row tuples), unlike the schema-based
    version above, which uses openpyxl's 1-based column numbering.
    """
    if not rows:
        return []
    data_start = find_data_start_row_from_rows(rows)
    sample_rows = rows[data_start:data_start + 60]
    max_col = max((len(r) for r in sample_rows), default=0)
    candidates = []
    for ci in range(max_col):
        vals = [r[ci] for r in sample_rows if ci < len(r) and r[ci] is not None]
        if len(vals) < 10:
            continue
        if vals and all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in vals):
            candidates.append(ci)
    return candidates


def find_probable_header_row(schema: dict, candidate_cols: list) -> int:
    """
    Among the header rows, find the one with the most non-empty values
    across the given candidate columns - the probable label row for those
    columns, regardless of whether any label matches a KNOWN pattern yet.

    Deliberately NOT detect_date_axis_row(): that function's fallback
    (return 1) fires exactly when NO header matches a known pattern - i.e.
    exactly the situation this needs to handle (a genuinely new format).
    Trusting that fallback here would silently point at the wrong row
    (often a title row) and miss every real candidate column's header text.
    """
    row_counts = {}
    col_schema_map = {c["col"]: c for c in schema.get("columns", [])}
    for col_1 in candidate_cols:
        cs = col_schema_map.get(col_1, {})
        for h in cs.get("header_stack", []):
            if str(h.get("value", "")).strip():
                row_counts[h["row"]] = row_counts.get(h["row"], 0) + 1
    if not row_counts:
        return 1
    return max(row_counts, key=row_counts.get)


def find_probable_header_row_from_rows(rows: list, candidate_cols: list) -> int:
    """Raw-rows equivalent of find_probable_header_row() - which of the
    first several rows has non-empty string values for the most candidate
    columns. 0-based row index, matching raw row tuples."""
    row_counts = {}
    for ri, row in enumerate(rows[:15]):
        count = sum(
            1 for ci in candidate_cols
            if ci < len(row) and isinstance(row[ci], str) and row[ci].strip()
        )
        if count:
            row_counts[ri] = count
    if not row_counts:
        return 0
    return max(row_counts, key=row_counts.get)


def build_new_pattern_prompt(shape_groups: dict, filename: str) -> str:
    """
    shape_groups: {shape_key: [(col_num, header_text), ...]} - a FEW
    examples per DISTINCT header shape (spanning the range seen, not just
    one), not every individual column. Showing the range matters: a single
    example like "202601 Units" is genuinely ambiguous (01 could be month
    or week) - only seeing a value like "202652" rules out month. AI's job
    is narrowed to correctly generalizing each distinct shape into a regex
    WITH capture groups identifying which part is the year/week/month/day -
    not to state what any one example means. Python computes the actual
    date for every matched column itself (see compute_date_from_match),
    using those capture groups against each column's OWN header text - so
    every column gets its own genuinely correct date, not a copy of
    whatever one example represented.
    """
    group_blocks = []
    for examples in shape_groups.values():
        lines = "\n".join(f"  - {repr(h)}" for _, h in examples)
        group_blocks.append(lines)
    header_lines = "\n\n".join(group_blocks)
    return f"""You are identifying whether column headers in a sales spreadsheet represent a date or sales period.

File: {filename}

These are examples of DISTINCT header shapes among columns that already contain
real sales-shaped numeric data (confirmed by Python), but don't match any known
date format. Each group below shows multiple examples of the SAME shape,
spanning the actual range of values seen - not just one instance:

{header_lines}

IMPORTANT - a common ambiguity: a 2-digit trailing number could be a MONTH
(01-12) or a WEEK NUMBER (01-53). Do not assume month by default. Look at the
actual range of examples shown for each shape - if any value exceeds 12, it
cannot be a month, and must be a week number (or something else entirely).
Only conclude "month" if every example you can see is consistent with a
12-month range AND you have another reason to believe it's monthly data.

For each shape that genuinely represents a sales date or period, provide a
generalized regex pattern WITH CAPTURE GROUPS marking which part is the year
and which part is the week (or month/day, if applicable) - Python will use
these groups to compute the correct date for every column that matches this
pattern, not just the examples shown. Do not just describe what one example
means; the year/week must be extractable from ANY matching header via the
capture groups you provide.

Respond with JSON only:
{{"resolutions": [
  {{"example_header": "202601 Units",
    "general_pattern": "^(\\\\d{{4}})(\\\\d{{2}})\\\\s*Units$",
    "capture_groups": {{"year": 1, "week": 2}},
    "resolution_method": "iso_year_week",
    "pattern_description": "YYYYWW Units"}}
]}}

resolution_method must be one of: "iso_year_week" (capture_groups: year, week),
"date_range" (already handled, shouldn't appear here), or "unknown" if the
format is a genuine date/period but doesn't fit a method Python can compute yet
(Python will still use it to identify which columns are date columns, but won't
be able to compute a specific date for them without further work).

If an example is NOT actually a date/period column (e.g. a product ID, UPC,
or a summary/total column), omit it entirely rather than guessing."""


async def resolve_unrecognized_dates(unresolved: list, filename: str, retailer: str | None) -> tuple[dict, dict]:
    """
    Escalates only the small residual (sales-shaped data + unrecognized
    header) to AI - and only ONE representative example per distinct header
    SHAPE, not every individual column. AI's job is narrowed to correctly
    generalizing each distinct shape into a regex WITH capture groups
    identifying year/week - it never states what any single example means.
    PYTHON then applies that pattern across the FULL unresolved list AND
    computes each matched column's own date from its OWN captured digits
    (see compute_date_from_match) - so a genuinely new format's actual
    per-column meaning is computed correctly for every column, not copied
    from whichever one example AI happened to look at.

    Before asking AI anything, checks EVERY unresolved header against ALL
    existing patterns - active AND pending_review, not just the active-only
    cache used for trusted matching. Without this, reprocessing the same (or
    a similarly-formatted) file re-escalates to AI and writes a fresh
    pending_review row every single run.

    Returns (resolved, trusted_dates):
      - resolved: {col: date_or_None} for EVERY column that matches a
        confirmed-genuine pattern, active or pending - used for column
        ENUMERATION (sales_cols_1based), same as before this change.
      - trusted_dates: {col: date} ONLY for columns backed by an
        ALREADY-ACTIVE pattern at resolution time - safe to persist into
        date_config["resolved_dates"] for ingestion to use directly,
        without recomputing anything or re-checking the pattern table.
        A column resolved via a BRAND-NEW pattern discovered THIS run is
        never in trusted_dates, even though its date IS computed correctly
        (see resolved) - that pattern is pending_review, and persisting its
        date as ingestion-ready would silently bypass the human approval
        gate. It only earns a spot in trusted_dates on a LATER run, once
        the pattern's been approved in the meantime.

    Validates the AI's own generalized pattern actually matches the example
    that produced it before trusting or applying it anywhere. Fails safe:
    any AI/parse error returns whatever was already resolved via existing
    patterns, and the caller's existing disqualify path still runs for the
    rest, exactly as it did before this feature existed.
    """
    if not unresolved:
        return {}, {}

    try:
        existing_patterns = await call_postgres(build_fetch_existing_patterns_for_dedup_sql())
    except Exception:
        existing_patterns = []

    def _match_existing(header: str) -> dict | None:
        for p in existing_patterns:
            try:
                if re.match(p["pattern_regex"], header.strip(), re.IGNORECASE):
                    return p
            except (re.error, TypeError, KeyError):
                continue
        return None

    resolved         = {}
    trusted_dates    = {}
    still_unresolved = []
    for col, header in unresolved:
        existing = _match_existing(header)
        if existing:
            m = re.match(existing["pattern_regex"], header.strip(), re.IGNORECASE)
            computed = compute_date_from_match(m, existing.get("resolution_rule", {})) if m else None
            resolved[col] = computed
            if computed and existing.get("status") == "active":
                trusted_dates[col] = computed
        else:
            still_unresolved.append((col, header))

    if not still_unresolved:
        return resolved, trusted_dates  # everything already covered by a previously-logged pattern - no AI call needed

    # Group by shape, keeping a FEW examples spanning the range - not just
    # the first one seen. A single example is often genuinely ambiguous
    # (e.g. "202601 Units": 01 could be month or week); seeing the column
    # with the lowest AND highest index within a shape (columns run in
    # temporal order in every real file seen so far) gives AI the range it
    # needs to tell month and week formats apart, rather than defaulting to
    # whichever is more common in general data (month).
    shape_groups = {}
    for col, header in still_unresolved:
        shape = normalize_header_shape(header)
        shape_groups.setdefault(shape, []).append((col, header))
    for shape, examples in shape_groups.items():
        examples.sort(key=lambda ch: ch[0])
        if len(examples) > 2:
            shape_groups[shape] = [examples[0], examples[-1]]

    try:
        text   = await call_ai(build_new_pattern_prompt(shape_groups, filename), label="new_date_pattern")
        clean  = parse_ai_response(text)
        result = json.loads(clean)
    except (json.JSONDecodeError, ValueError):
        return resolved, trusted_dates
    except Exception:
        return resolved, trusted_dates

    written_patterns = set()

    for r in result.get("resolutions", []) if isinstance(result, dict) else []:
        pattern           = r.get("general_pattern")
        example           = r.get("example_header", "")
        resolution_rule   = {
            "method": r.get("resolution_method", "unknown"),
            "capture_groups": r.get("capture_groups", {}),
        }
        if not pattern:
            continue
        try:
            compiled = re.compile(pattern, re.IGNORECASE)
            if example and not compiled.match(example.strip()):
                continue  # AI's own pattern doesn't even match its own example - reject
        except re.error:
            continue

        # PYTHON applies the confirmed pattern across the REMAINING
        # unresolved headers, computing EACH column's own date from ITS OWN
        # match - never copying one example's meaning onto a different column.
        # Newly-discovered THIS RUN -> pending_review -> resolved[] only,
        # never trusted_dates (see docstring).
        matched_any = False
        for col, header in still_unresolved:
            m = compiled.match(header.strip())
            if m:
                resolved[col] = compute_date_from_match(m, resolution_rule)
                matched_any = True

        if matched_any and pattern not in written_patterns:
            written_patterns.add(pattern)
            try:
                await call_postgres(build_insert_date_pattern_sql(
                    retailer, pattern, r.get("pattern_description", "AI-discovered"),
                    resolution_rule, example, filename,
                ))
            except Exception:
                pass  # write-back failing shouldn't block resolving THIS file's columns

    return resolved, trusted_dates


def build_column_classify_prompt(schema: dict, filename: str,
                                  date_cols: list, postgres_matched: set) -> str:
    """
    Pass 2 - AI classifies non-date columns only.
    Small focused prompt - only the product/inventory/metadata columns.
    """
    date_col_set = set(date_cols)
    non_date_cols = [
        c for c in schema.get("columns", [])
        if c["col"] not in date_col_set
        and (c.get("sample_non_zero") or c.get("header_stack"))
    ]

    # Annotate with postgres match info
    for col in non_date_cols:
        col["postgres_matched"] = any(
            str(v).upper() in postgres_matched
            for v in col.get("sample_non_zero", [])
        )
        col["embedded_postgres_matched"] = any(
            ext.get("sku", "").upper() in postgres_matched
            for ext in col.get("embedded_sku_extractions", [])
        )

    cols_json = json.dumps(non_date_cols, indent=2)

    return f"""You are classifying the non-sales columns of a retail sales spreadsheet.

File: {filename}
Sheet: {schema["sheet_name"]}

The sales date columns have already been identified. Classify only the columns below.
You MUST return a classification for EVERY column listed — do not skip any.

Each column has:
- header_stack: header values with row numbers (top-down)
- data_types: Python types found in data rows
- sample_non_zero: up to 5 non-zero sample values
- pct_zero: fraction of data rows that are zero/null
- postgres_matched: true if sample values matched the supplier inventory database
- embedded_sku_extractions: supplier SKUs extracted from text values
- embedded_postgres_matched: true if extracted SKUs matched the supplier database

Classify each column as one of:
- retailer_sku: retailer's product identifier (WIC#, DPCI, SKU Number, Item# etc)
- supplier_sku: supplier's internal part number or style code
- description: product name/description text (may contain embedded supplier SKU)
- cost: unit cost or wholesale price
- retail_price: retail selling price
- inventory: the single best current total on-hand quantity to ingest - pick exactly one per sheet
- open_order: open purchase order quantity (even if currently zero/empty)
- other: everything else - DC/store inventory sub-components, prior snapshots, YTD totals, percentages, business unit codes, store counts, etc

For inventory: if a sheet has sub-components (DC Inv, Store Inv) alongside a total (Total Inv), classify only the total as inventory. If snapshots are dated, use the most recent total. Only one column per sheet should be classified as inventory.
- postgres_matched strongly suggests retailer_sku or supplier_sku
- embedded_postgres_matched means has_embedded_supplier_sku = true

COLUMNS ({len(non_date_cols)} total - classify all of them):
{cols_json}

Respond with JSON only:
{{
  "columns": [
    {{"col": 1, "classification": "retailer_sku", "confidence": "high", "reason": "one sentence", "has_embedded_supplier_sku": false}},
    ...
  ]
}}"""

# ---------------------------------------------
# FILE TYPE DETECTION
# ---------------------------------------------

# Keywords that identify known non-sales file types.
# Maps keyword (lowercase, found in filename or sheet names) -> audit status
KNOWN_FILE_TYPES = {
    "on hand inventory alpha": "inventory_report",  # primary match - most specific first
    "on hand":                 "inventory_report",  # fallback
}

def detect_known_file_type(filename: str, sheet_names: list) -> str | None:
    """
    Check filename and sheet names against known non-sales file type keywords.
    Returns the audit status to assign, or None if unrecognized.
    """
    haystack = ((filename or "") + " " + " ".join(sheet_names)).lower()
    for keyword, status in KNOWN_FILE_TYPES.items():
        if keyword in haystack:
            return status
    return None


# ---------------------------------------------
# PIPELINE STAGES
# ---------------------------------------------

async def _qualify_one_sheet(session: dict, sheet_name: str, rows: list):
    """
    Everything stage_qualify used to do inline, per sheet, in a sequential
    for-loop - now a standalone coroutine so multiple sheets can run
    concurrently via asyncio.gather instead of one after another.

    Safe to parallelize: this function's logic depends only on this
    sheet's own rows/signals - it never reads another sheet's
    qualify_results, qualified_sheets, or any other sheet-keyed state.
    Every write below is to a key namespaced by sheet_name (or an append),
    so concurrent coroutines writing different sheets' results can't
    clobber each other under asyncio's cooperative single-threaded model.
    """
    signals = extract_qualify_signals(rows, sheet_name, session["filename"])

    # No known pattern matched anything for this sheet's date axis.
    # Before trusting that as "no sales data" evidence (which would
    # likely disqualify the sheet here, PERMANENTLY - stage_qualify runs
    # before stage_schema_classify, so if this sheet is rejected now, the
    # schema-level escalation built for this same purpose is never
    # reached), check whether sales-shaped numeric data exists under
    # headers that just aren't recognized yet, and escalate that
    # residual to AI right here, at qualify time.
    if signals.get("date_col_count", 0) == 0:
        shaped_cols = find_sales_shaped_columns_from_rows(rows)
        if shaped_cols:
            header_row = find_probable_header_row_from_rows(rows, shaped_cols)
            unresolved = []
            for ci in shaped_cols:
                header_val = ""
                if header_row < len(rows) and ci < len(rows[header_row]) and rows[header_row][ci] is not None:
                    header_val = str(rows[header_row][ci]).strip()
                if header_val and not match_known_patterns(header_val):
                    unresolved.append((ci, header_val))

            if unresolved:
                newly_resolved, trusted_dates = await resolve_unrecognized_dates(
                    unresolved, session["filename"], session.get("retailer")
                )
                if newly_resolved:
                    signals["date_col_count"] = len(newly_resolved)
                    # Store so stage_schema_classify can reuse this exact
                    # resolution later instead of re-escalating to AI a
                    # second time for the same header shape in this file.
                    session.setdefault("_qualify_resolved_dates", {})[sheet_name] = newly_resolved
                    # Only active-pattern-backed dates get persisted for
                    # ingestion to use directly - see resolve_unrecognized_dates.
                    if trusted_dates:
                        session.setdefault("_trusted_resolved_dates", {}).setdefault(
                            sheet_name, {}
                        ).update(trusted_dates)
                    session.setdefault("flags", {})["new_date_pattern_discovered"] = (
                        f"{len(newly_resolved)} column(s) resolved via AI at qualify stage "
                        f"for sheet '{sheet_name}' - pattern(s) written as pending_review"
                    )

    # Deterministic fast paths
    if signals["dominant_type"] == "float_above_1":
        session["qualify_results"][sheet_name] = {
            "disqualified": True,
            "reason": "Crosshair values predominantly decimals above 1 - dollar revenue",
            "source": "python",
        }
        return
    if signals["dominant_type"] == "float_0_to_1":
        session["qualify_results"][sheet_name] = {
            "disqualified": True,
            "reason": "Crosshair values predominantly between 0 and 1 - percentage metrics",
            "source": "python",
        }
        return

    # AI call - direct, no webhook
    try:
        text   = await call_ai(build_qualify_prompt(signals), label="qualify")
        clean  = parse_ai_response(text)
        verdict = json.loads(clean)
    except (json.JSONDecodeError, ValueError) as e:
        verdict = {"disqualified": None, "reason": f"Parse error: {e}"}
    except Exception as e:
        verdict = {"disqualified": None, "reason": f"AI error: {e}"}
    verdict["source"] = "ai"
    session["qualify_results"][sheet_name] = verdict


async def stage_qualify(session_id: str):
    """Stage 1 - qualify each sheet. Sheets are independent, so they run
    concurrently rather than one after another - see _qualify_one_sheet."""
    session = _sessions[session_id]
    session["stage"]  = "qualifying"
    session["status"] = "running"

    # Load the approved date-format pattern library once per run. Retailer is
    # not known yet at this stage, so this loads ALL active patterns, any
    # retailer - evidence a header shape is a real, previously-approved date
    # axis, regardless of whose file this turns out to be.
    await load_date_patterns()

    await asyncio.gather(*[
        _qualify_one_sheet(session, sheet_name, rows)
        for sheet_name, rows in session["_sheets"].items()
    ])

    await advance_from_qualify(session_id)


async def advance_from_qualify(session_id: str):
    """After all qualify jobs done - filter qualified sheets, advance to grid location."""
    session = _sessions[session_id]
    results = session["qualify_results"]

    qualified = [
        name for name, r in results.items()
        if not r.get("disqualified")
    ]

    session["qualified_sheets"] = qualified

    if not qualified:
        file_audit_id = session.get("file_audit_id")
        # Check if this is a known non-sales file type before marking rejected
        known_type = detect_known_file_type(
            session.get("filename", ""),
            list(session.get("_sheets", {}).keys())
        )
        audit_status = known_type or "rejected"
        if file_audit_id:
            try:
                sql = build_update_file_audit_full_sql(
                    file_audit_id,
                    {"status": "no_sales_data", "qualify_results": results},
                    None, audit_status, None,
                    file_hash=session.get("file_hash"),
                    filename=session.get("filename"),
                )
                await call_postgres(sql)
            except Exception as e:
                session.setdefault("errors", []).append(f"Failed to write {audit_status} status: {e}")
        session["stage"]  = "complete"
        session["status"] = "complete"
        session["result"] = {"status": audit_status, "qualify_results": results}
        return

    await stage_locate_grid(session_id)


async def stage_locate_grid(session_id: str):
    """
    Stage 2 - pure Python schema extraction.
    Opens workbook with full fidelity (merges + borders).
    Extracts per-column schema for all qualified sheets.
    Then advances to Postgres SKU lookup.
    """
    session  = _sessions[session_id]
    session["stage"]  = "locating"
    session["status"] = "running"

    # Need full workbook (not read-only) for merge + border detection
    wb_full = openpyxl.load_workbook(io.BytesIO(session["_raw_data"]), data_only=True)
    session["_schemas"] = {}

    for sheet_name in session["qualified_sheets"]:
        if sheet_name not in wb_full.sheetnames:
            session["grid"][sheet_name] = {"error": "Sheet not found in workbook"}
            continue
        ws     = wb_full[sheet_name]
        schema = extract_column_schema(ws, sheet_name)

        # Embedded SKU detection runs after schema_classify when we know sales_cols
        schema["embedded_sku"] = []

        session["_schemas"][sheet_name] = schema

    await stage_postgres_sku_lookup(session_id)





async def stage_postgres_sku_lookup(session_id: str):
    """Stage 2b - Postgres SKU lookup on candidate column values from all sheets."""
    session = _sessions[session_id]
    session["stage"]  = "identifying"
    session["status"] = "running"

    col_candidates = {}
    for sheet_name in session["qualified_sheets"]:
        schema     = session.get("_schemas", {}).get(sheet_name, {})
        rows       = session["_sheets"].get(sheet_name, [])
        data_start = schema.get("data_start_row", 1) - 1

        for col in schema.get("columns", []):
            for val in col.get("sample_non_zero", []):
                s = str(val).strip()
                if s:
                    col_candidates.setdefault(s, set()).add(col["col"])

            if "string" in col.get("data_types", []):
                col_0    = col["col"] - 1
                str_vals = [
                    str(rows[r][col_0]).strip()
                    for r in range(data_start, min(data_start + 30, len(rows)))
                    if col_0 < len(rows[r]) and isinstance(rows[r][col_0], str)
                    and rows[r][col_0].strip()
                ]
                for val in str_vals:
                    for _, pattern_re in EMBEDDED_PATTERNS:
                        m = pattern_re.match(val)
                        if m:
                            sku = m.group(2).strip() if pattern_re.pattern.startswith(r'^(.{3') else m.group(1).strip()
                            if sku:
                                col_candidates.setdefault(sku, set()).add(col["col"])
                            break

    session["_col_candidates"] = {k: list(v) for k, v in col_candidates.items()}

    try:
        sql, _ = build_sku_lookup_sql(sorted(col_candidates.keys()))
        matches = await call_postgres(sql)
        session["postgres_results"] = {"matches": matches}
    except Exception as e:
        session["postgres_results"] = {"matches": [], "error": str(e)}

    await stage_schema_classify(session_id)


async def _schema_classify_one_sheet(session: dict, sheet_name: str, matched_values: set):
    """
    Everything stage_schema_classify used to do inline, per sheet, in a
    sequential for-loop - now a standalone coroutine so multiple sheets
    can run concurrently via asyncio.gather instead of one after another.

    Safe to parallelize: depends only on this sheet's own schema/rows and
    on _matched_values (read-only, set once before any sheet is processed).
    Every write is namespaced by sheet_name, or - for the one shared-list
    mutation (session["qualified_sheets"], when a sheet turns out to have
    no date columns after all) - a same-value-removal filter with no
    intervening await, which is safe under asyncio's cooperative
    single-threaded model regardless of how multiple sheets' coroutines
    happen to interleave.
    """
    schema = session.get("_schemas", {}).get(sheet_name, {})
    if not schema:
        session["grid"][sheet_name]           = {"error": "No schema available"}
        session["column_mapping"][sheet_name] = []
        return

    for col in schema.get("columns", []):
        col["postgres_matched"] = any(
            str(v).upper() in matched_values for v in col.get("sample_non_zero", [])
        )
        embedded_matched = any(
            ext.get("sku", "").upper() in matched_values
            for ext in col.get("embedded_sku_extractions", [])
        )
        col["embedded_postgres_matched"] = embedded_matched
        if embedded_matched:
            col["has_embedded_supplier_sku"] = True

    # Pass 1 - AI identifies date schema pattern
    try:
        text        = await call_ai(build_date_schema_prompt(schema, session["filename"]), label="date_schema")
        clean       = parse_ai_response(text)
        date_schema = json.loads(clean)
    except (json.JSONDecodeError, ValueError) as e:
        date_schema = {"error": f"Parse error: {e}"}
    except Exception as e:
        date_schema = {"error": f"AI error: {e}"}

    if "error" in date_schema:
        session["grid"][sheet_name]           = {"error": date_schema["error"]}
        session["column_mapping"][sheet_name] = []
        return

    # Python enumerates all sales date columns from the pattern
    sales_cols_1based = find_sales_cols_from_schema(schema, date_schema)

    # Nothing matched a known, approved pattern - before disqualifying,
    # check whether Python-confirmed sales-shaped DATA exists under
    # headers that just aren't recognized yet, and escalate only that
    # residual to AI. The AI's resolution is used to finish processing
    # THIS file now; the generalized pattern it writes back is stored as
    # pending_review and does NOT become a trusted match for any other
    # file (or even a re-run of this same enumeration) until a human
    # approves it - matching how new retailers are already handled.
    if not sales_cols_1based:
        # Reuse a resolution already made at qualify time for this exact
        # sheet, if one exists - avoids a second AI call and a second
        # pending_review pattern write for the same header shape within
        # one file's processing. Qualify-time keys are 0-based (raw row
        # tuples); this stage's are 1-based (openpyxl/schema convention).
        qualify_resolved = session.get("_qualify_resolved_dates", {}).get(sheet_name)
        if qualify_resolved:
            sales_cols_1based = sorted(c0 + 1 for c0 in qualify_resolved.keys())

    if not sales_cols_1based:
        shaped_cols = find_sales_shaped_columns(schema)
        if shaped_cols:
            # NOT detect_date_axis_row() here - its fallback (row 1) fires
            # exactly when no header matches a known pattern, which is
            # precisely this situation. find_probable_header_row() finds
            # the label row structurally, independent of pattern matching.
            header_row_probe      = find_probable_header_row(schema, shaped_cols)
            col_schema_map_probe = {c["col"]: c for c in schema.get("columns", [])}
            unresolved = []
            for col_1 in shaped_cols:
                cs = col_schema_map_probe.get(col_1, {})
                header_val = ""
                for h in cs.get("header_stack", []):
                    if h["row"] == header_row_probe:
                        header_val = h["value"]
                        break
                if header_val and not match_known_patterns(header_val):
                    unresolved.append((col_1, header_val))

            if unresolved:
                newly_resolved, trusted_dates = await resolve_unrecognized_dates(
                    unresolved, session["filename"], session.get("retailer")
                )
                if newly_resolved:
                    sales_cols_1based = sorted(newly_resolved.keys())
                    # date_cols/resolved_dates use 0-based indexing
                    # (matching sales_cols_0based below) - this
                    # escalation path works in 1-based (schema/openpyxl
                    # convention), so convert before merging.
                    if trusted_dates:
                        session.setdefault("_trusted_resolved_dates", {}).setdefault(
                            sheet_name, {}
                        ).update({c1 - 1: d for c1, d in trusted_dates.items()})
                    session.setdefault("flags", {})["new_date_pattern_discovered"] = (
                        f"{len(newly_resolved)} column(s) resolved via AI for this file - "
                        f"pattern(s) written to date_format_patterns as pending_review, "
                        f"needs human approval before automatic reuse on other files"
                    )

    sales_cols_0based = [c - 1 for c in sales_cols_1based]

    session.setdefault("_date_schemas", {})[sheet_name] = date_schema
    session.setdefault("_sales_cols",   {})[sheet_name] = sales_cols_0based

    # Pass 2 - AI classifies non-date columns
    try:
        text   = await call_ai(build_column_classify_prompt(schema, session["filename"],
                                                              sales_cols_1based, matched_values), label="column_classify")
        clean  = parse_ai_response(text)
        result = json.loads(clean)
    except (json.JSONDecodeError, ValueError) as e:
        result = {"error": f"Parse error: {e}"}
    except Exception as e:
        result = {"error": f"AI error: {e}"}

    if "error" in result:
        session["grid"][sheet_name]           = {"error": result["error"]}
        session["column_mapping"][sheet_name] = []
        return

    col_results   = result.get("columns", [])
    col_results   = [{**c, "col": c["col"] - 1} for c in col_results]

    data_start    = schema["data_start_row"]               # Python
    # NOT detect_date_axis_row() here either - same reason as the
    # escalation path above: its fallback (row 1) fires precisely when
    # headers don't match an ACTIVE pattern, which is exactly the case
    # for a format discovered THIS run (still pending_review) or one
    # whose header-row detection otherwise fails. That silently pointed
    # ingestion at the wrong row (often a title row) for any file using
    # a newly-discovered format - build_date_map would find nothing but
    # None values there, sales_rows would come back 0, and nothing
    # would report this as an error. find_probable_header_row() finds
    # the label row structurally, independent of pattern status,
    # working correctly for both known and newly-discovered formats.
    date_axis_row = find_probable_header_row(schema, sales_cols_1based)  # Python (1-based)
    rows_data     = session["_sheets"].get(sheet_name, [])
    axis_row_0    = date_axis_row - 1                      # 0-based

    # Detect format and year flags from actual header values
    axis_row_vals = [
        rows_data[axis_row_0][c]
        for c in sales_cols_0based
        if axis_row_0 < len(rows_data) and c < len(rows_data[axis_row_0])
        and rows_data[axis_row_0][c] is not None
    ]
    formats = set(classify_cell(v) for v in axis_row_vals)
    date_fmt = list(formats)[0] if len(formats) == 1 else "mixed"

    year_present  = date_fmt == "datetime" or any(
        YEAR_RE.search(str(v)) for v in axis_row_vals
    )

    col_schema_map = {c["col"]: c for c in schema["columns"]}
    sample_vals    = []
    for col_0 in sorted(sales_cols_0based)[:8]:
        col_1 = col_0 + 1
        cs    = col_schema_map.get(col_1, {})
        for h in cs.get("header_stack", []):
            if any(ch.isdigit() for ch in h["value"]):
                sample_vals.append(h["value"])
                break

    # Also grab the last date column value for file_set_key calculation
    last_sample_val = None
    if sales_cols_0based:
        last_col_0 = sorted(sales_cols_0based)[-1]
        last_col_1 = last_col_0 + 1
        cs = col_schema_map.get(last_col_1, {})
        for h in cs.get("header_stack", []):
            if any(ch.isdigit() for ch in h["value"]):
                last_sample_val = h["value"]
                break

    # Find last ACTIVE column - last date column with at least one non-zero data value
    # This is the true latest week for file_set_key, not the last column in the sheet
    last_active_sample_val = None
    data_start_0 = data_start - 1  # 0-based
    for col_0 in sorted(sales_cols_0based, reverse=True):
        has_data = any(
            rows_data[r][col_0] not in (None, 0, "", " ")
            for r in range(data_start_0, min(data_start_0 + 200, len(rows_data)))
            if col_0 < len(rows_data[r])
        )
        if has_data:
            col_1 = col_0 + 1
            cs = col_schema_map.get(col_1, {})
            for h in cs.get("header_stack", []):
                if any(ch.isdigit() for ch in h["value"]):
                    last_active_sample_val = h["value"]
                    break
            break

    year_anchors = [
        {"source": "filename", "value": m.group(1)}
        for m in YEAR_RE.finditer(session["filename"])
    ]

    # Detect year boundary from actual month sequence
    # A boundary exists only when December is followed by January (month drops)
    # A full-year file (Jan->Dec) contains both months but has no boundary
    month_seq      = extract_month_sequence(rows_data[axis_row_0] if axis_row_0 < len(rows_data) else [], sales_cols_0based)
    year_boundary  = False
    prev_m = None
    for m in month_seq:
        if m is None:
            continue
        if prev_m is not None and prev_m == 12 and m == 1:
            year_boundary = True
            break
        prev_m = m

    date_axis = {
        "row":                    axis_row_0,
        "col_count":              len(sales_cols_0based),
        "date_col_count":         len(sales_cols_0based),
        "cols":                   sales_cols_0based,
        "sample_values":          sample_vals,
        "last_sample_val":        last_sample_val,
        "last_active_sample_val": last_active_sample_val,
        "format":                 date_fmt,
        "year_present":           year_present,
        "interleaved_empty_cols": False,
        "year_boundary_detected": year_boundary,
    }

    # Store month_sequence if boundary detected
    if year_boundary:
        date_axis["month_sequence"] = month_seq

    rows     = session["_sheets"].get(sheet_name, [])
    embedded = detect_embedded_sku(rows, sales_cols_0based, data_start - 1)

    embedded_by_col = {e["col"]: e for e in embedded}
    for col in schema.get("columns", []):
        col_0 = col["col"] - 1
        if col_0 in embedded_by_col:
            col["embedded_sku_extractions"] = embedded_by_col[col_0]["extractions"]
            col["has_embedded_supplier_sku"] = True

    sku_candidates = []
    for c in col_results:
        col_0 = c["col"]
        col_1 = col_0 + 1
        cs    = col_schema_map.get(col_1, {})
        pre_strings = [
            {"row": h["row"] - 1, "value": h["value"]}
            for h in cs.get("header_stack", [])
        ]
        col_vals = [
            rows[r][col_0]
            for r in range(data_start - 1, min(data_start + 49, len(rows)))
            if col_0 < len(rows[r]) and rows[r][col_0] is not None
        ]
        types   = {}
        for v in col_vals:
            t = classify_cell(v)
            types[t] = types.get(t, 0) + 1
        dominant = max(types, key=types.get) if types else "string"
        sku_candidates.append({
            "col":              col_0,
            "pre_data_strings": pre_strings,
            "dominant_type":    dominant,
            "type_distribution": types,
            "fill_rate":        round(len(col_vals) / 50, 2),
            "sample_values":    [str(v) for v in col_vals[:5]],
        })

    session["grid"][sheet_name] = {
        "date_axis":      date_axis,
        "data_start_row": data_start - 1,
        "sku_candidates": sku_candidates,
        "embedded_sku":   embedded,
        "year_anchors":   year_anchors[:6],
    }

    def _has_inventory_keyword(col_0based: int) -> bool:
        """Does this column's own header text actually contain a real
        inventory keyword? Used to gate the "inventory" classification -
        AI's free-form reasoning about what a column "represents" isn't
        reliable enough on its own for a label that determines which
        TABLE data gets written into (see INVENTORY_LABELS's docstring
        for the real production case that showed this: the same column
        classified two different ways on two runs of the same file)."""
        cs = col_schema_map.get(col_0based + 1, {})
        header_text = " ".join(
            str(h.get("value", "")) for h in cs.get("header_stack", []) if h.get("value")
        ).lower()
        return any(kw in header_text for kw in INVENTORY_LABELS)

    column_mapping_list = []
    for c in col_results:
        if c.get("classification") == "sales_date":
            continue
        classification = c.get("classification")
        if classification == "inventory" and not _has_inventory_keyword(c["col"]):
            # AI said inventory, but the header itself has no real
            # inventory keyword - don't trust the label, fall back to
            # the safe default rather than risk sales data landing in
            # the inventory table.
            classification = "other"
        column_mapping_list.append({
            "col":                  c["col"],
            "classification":       classification,
            "confidence":           c.get("confidence", "high"),
            "reason":               c.get("reason", ""),
            "has_embedded_supplier_sku": (
                c.get("has_embedded_supplier_sku", False) or
                col_schema_map.get(c["col"] + 1, {}).get("has_embedded_supplier_sku", False) or
                col_schema_map.get(c["col"] + 1, {}).get("embedded_postgres_matched", False)
            ),
        })
    session["column_mapping"][sheet_name] = column_mapping_list

    # Post-schema: disqualify if no sales date columns found
    if len(sales_cols_0based) == 0:
        session["qualified_sheets"] = [s for s in session["qualified_sheets"] if s != sheet_name]
        session["qualify_results"][sheet_name] = {
            "disqualified": True,
            "reason": "No weekly sales date columns identified after full layout analysis",
            "source": "post_schema",
        }
        session["grid"].pop(sheet_name, None)
        session["column_mapping"].pop(sheet_name, None)


async def stage_schema_classify(session_id: str):
    """Stage 2c - Pass 1: AI identifies date column pattern per sheet.
    Sheets are independent, so they run concurrently rather than one
    after another - see _schema_classify_one_sheet."""
    session = _sessions[session_id]
    session["stage"]  = "locating"
    session["status"] = "running"

    pg_matches     = session.get("postgres_results", {}).get("matches", [])
    matched_values = set()
    for row in pg_matches:
        for field in ("inventory_sku", "base_variant", "base_model"):
            v = row.get(field, "")
            if v:
                matched_values.add(v.upper())
    session["_matched_values"] = matched_values

    await asyncio.gather(*[
        _schema_classify_one_sheet(session, sheet_name, matched_values)
        for sheet_name in session["qualified_sheets"]
    ])

    if not session["qualified_sheets"]:
        file_audit_id = session.get("file_audit_id")
        known_type = detect_known_file_type(
            session.get("filename", ""),
            list(session.get("_sheets", {}).keys())
        )
        audit_status = known_type or "rejected"
        if file_audit_id:
            try:
                sql = build_update_file_audit_full_sql(
                    file_audit_id,
                    {"status": "no_sales_data", "qualify_results": session["qualify_results"]},
                    None, audit_status, None,
                    file_hash=session.get("file_hash"),
                    filename=session.get("filename"),
                )
                await call_postgres(sql)
            except Exception as e:
                session.setdefault("errors", []).append(f"Failed to write {audit_status} status: {e}")
        session["stage"]  = "complete"
        session["status"] = "complete"
        session["result"] = {"status": audit_status, "qualify_results": session["qualify_results"]}
        return

    await stage_identify_retailer(session_id)


async def stage_identify_retailer(session_id: str):
    """Stage 3 - identify retailer by querying retailer_sku_map."""
    session = _sessions[session_id]
    session["stage"]  = "identifying_retailer"
    session["status"] = "running"

    retailer_sku_candidates = set()
    for sheet_name in session["qualified_sheets"]:
        mapping    = session["column_mapping"].get(sheet_name, [])
        cols       = mapping if isinstance(mapping, list) else mapping.get("columns", [])
        grid       = session["grid"].get(sheet_name, {})
        rows       = session["_sheets"].get(sheet_name, [])
        data_start = grid.get("data_start_row", 0)
        for col_info in cols:
            if col_info.get("classification") == "retailer_sku":
                col_idx = col_info.get("col")
                for row in rows[data_start:]:
                    if col_idx < len(row) and row[col_idx] is not None:
                        v = str(row[col_idx]).strip()
                        if v:
                            retailer_sku_candidates.add(v)

    if retailer_sku_candidates:
        try:
            sql  = build_retailer_identify_sql(sorted(retailer_sku_candidates))
            rows = await call_postgres(sql)
            if rows and int(rows[0].get("matches", 0)) >= 3:
                session["retailer"] = rows[0].get("retailer", "").lower()
                session["flags"]["retailer_identification"] = (
                    f"confirmed: {session['retailer']} ({rows[0].get('matches')} matches)"
                )
                await stage_date_config(session_id)
                return
        except Exception as e:
            session["errors"].append(f"Retailer Postgres lookup failed: {e}")

    await stage_identify_retailer_ai(session_id)


async def stage_identify_retailer_ai(session_id: str):
    """Fallback - identify retailer via AI using filename/sheet/header clues."""
    session = _sessions[session_id]
    session["status"] = "running"

    header_strings = []
    for sheet_name in session["qualified_sheets"]:
        grid = session["grid"].get(sheet_name, {})
        for col in grid.get("sku_candidates", []):
            for s in col.get("pre_data_strings", []):
                v = s.get("value", "")
                if v and v not in header_strings:
                    header_strings.append(v)

    try:
        text   = await call_ai(build_retailer_identify_prompt(
            session["filename"], session["qualified_sheets"], header_strings
        ), label="retailer_identify")
        clean  = parse_ai_response(text)
        result = json.loads(clean)
    except (json.JSONDecodeError, ValueError) as e:
        result = {"retailer": None, "confidence": None, "reason": f"Parse error: {e}"}
    except Exception as e:
        result = {"retailer": None, "confidence": None, "reason": f"AI error: {e}"}

    session["retailer"] = result.get("retailer", "").lower() if result.get("retailer") else None
    session["flags"]["retailer_identification"] = (
        f"ai_confirmed: {session['retailer']} ({result.get('confidence', 'unknown')} confidence)"
        if session["retailer"]
        else f"ai_unconfirmed: {result.get('reason', 'unknown')}"
    )
    await stage_date_config(session_id)


def _finalize_date_config(session: dict, sheet_name: str, dc: dict) -> None:
    """
    Computes everything ingestion needs to resolve dates, so ingestion
    never has to figure anything out itself - only read this.

    Two things get computed here, both previously computed independently
    (and inconsistently) by ingestion.py at ingestion time:

    1. resolved_dates: the COMPLETE col_idx -> date mapping for every date
       column, not just ones from an AI-discovered, already-active pattern.
       Calls the same build_date_map() used to live only in ingestion.py,
       moved to shared.py so discovery can call it once and persist the
       answer. Stored at session["_final_resolved_dates"][sheet_name] -
       NOT inside dc/date_config - since it now lives in its own dedicated
       file_audit.resolved_dates column, kept in exactly one place rather
       than duplicated between a JSON sub-field and a top-level column.
       dc still carries a SEED value (from the qualify/escalation stages'
       already-trusted subset) for build_date_map's own internal fast path,
       but that seed is consumed here, not re-persisted back into dc.

    2. inventory_as_of_date: the single most-recent resolvable date found
       in the inventory column's header rows, if this sheet has one -
       previously scanned by ingestion at ingestion time using its own
       _resolve_date() calls. This one stays in date_config since it's a
       single scalar, not the thing that caused today's duplication problem.

    Safe to call even if something's missing (no inventory column, no
    rows) - those cases just leave the corresponding field empty/unset
    rather than raising.
    """
    rows = session.get("_sheets", {}).get(sheet_name, [])
    axis_row_0 = dc.get("date_axis_row", 0)
    header_row = rows[axis_row_0] if axis_row_0 < len(rows) else ()

    try:
        complete_map = build_date_map(header_row, dc.get("date_cols", []), dc)
        session.setdefault("_final_resolved_dates", {})[sheet_name] = {
            str(k): v.isoformat() for k, v in complete_map.items()
        }
    except Exception as e:
        session.setdefault("_final_resolved_dates", {}).setdefault(sheet_name, {})
        session.setdefault("errors", []).append(f"resolved_dates computation failed for '{sheet_name}': {e}")
    # This dict was only ever a seed for build_date_map's fast path - the
    # complete result lives at the session level now, not back in dc.
    dc.pop("resolved_dates", None)

    inventory_col = next(
        (c.get("col") for c in session.get("column_mapping", {}).get(sheet_name, [])
         if c.get("classification") == "inventory"),
        None,
    )
    if inventory_col is not None:
        data_start = dc.get("data_start_row", 1)
        year_value = dc.get("year_value", date.today().year)
        latest = None
        for hrow_idx in range(min(data_start, len(rows))):
            if inventory_col < len(rows[hrow_idx]) and rows[hrow_idx][inventory_col]:
                try:
                    parsed = resolve_date_header(str(rows[hrow_idx][inventory_col]).strip(), year_value)
                except Exception:
                    parsed = None
                if parsed and (latest is None or str(parsed) > latest):
                    latest = str(parsed)
        if latest:
            dc["inventory_as_of_date"] = latest


async def _date_config_one_sheet(session: dict, sheet_name: str):
    """
    Everything stage_date_config used to do inline, per sheet, in a
    sequential for-loop - now a standalone coroutine so multiple sheets
    can run concurrently via asyncio.gather instead of one after another.

    Safe to parallelize: cross_sheet_anchors below only READS other
    sheets' year_anchors, which are fully populated by stage_schema_classify
    (a prior stage that has completely finished, for every sheet, before
    this stage even begins) - nothing in this stage writes year_anchors,
    so there's no race between sheets reading each other's data here.
    """
    grid      = session["grid"].get(sheet_name, {})
    date_axis = grid.get("date_axis", {})
    anchors   = grid.get("year_anchors", [])

    if not date_axis:
        session["date_config"][sheet_name] = {"error": "No date axis"}
        return

    if date_axis.get("year_present"):
        year_value = None
        for a in anchors:
            y = str(a.get("value", ""))[:4]
            if y.isdigit():
                year_value = int(y)
                break
        if not year_value:
            for sample in date_axis.get("sample_values", []):
                parts = str(sample).replace("-", "/").split("/")
                if len(parts) >= 3:
                    y = parts[-1].strip()[:4]
                    if y.isdigit():
                        yr = int(y)
                        year_value = 2000 + yr if yr < 100 else yr
                        break
        year_boundary = date_axis.get("year_boundary_detected", False)
        month_seq     = date_axis.get("month_sequence", [])
        year_start    = compute_year_start(month_seq, year_value) if year_boundary and month_seq else year_value
        session["date_config"][sheet_name] = {
            "date_format":             date_axis["format"],
            "year_present":            True,
            "year_value":              year_value,
            "year_start":              year_start,
            "year_inference_strategy": "embedded_in_dates",
            "year_boundary_detected":  year_boundary,
            "normalize_to":            "week_ending_saturday",
            "source":                  "python",
            "date_axis_row":           date_axis.get("row", 0),
            "date_cols":               date_axis.get("cols", []),
            "data_start_row":          grid.get("data_start_row", 1),
            # Seed with whatever the qualify/escalation stages already
            # trusted - _finalize_date_config below computes the
            # COMPLETE map (every date column, not just this subset)
            # and overwrites this with the full result.
            "resolved_dates": {
                str(k): v for k, v in
                session.get("_trusted_resolved_dates", {}).get(sheet_name, {}).items()
            },
        }
        _finalize_date_config(session, sheet_name, session["date_config"][sheet_name])
        return

    cross_sheet_anchors = [
        a for other in session["qualified_sheets"] if other != sheet_name
        for a in session["grid"].get(other, {}).get("year_anchors", [])
    ]

    year_boundary  = date_axis.get("year_boundary_detected", False)
    month_sequence = date_axis.get("month_sequence", [])

    try:
        text   = await call_ai(build_date_prompt(date_axis, anchors, sheet_name,
                                                  filename=session["filename"],
                                                  cross_sheet_anchors=cross_sheet_anchors), label="date_config")
        clean  = parse_ai_response(text)
        result = json.loads(clean)
    except (json.JSONDecodeError, ValueError) as e:
        result = {"error": f"Parse error: {e}"}
    except Exception as e:
        result = {"error": f"AI error: {e}"}

    year_value = result.get("year_value", date.today().year)
    year_start = compute_year_start(month_sequence, year_value) if year_boundary and month_sequence else year_value

    result["normalize_to"]           = "week_ending_saturday"
    result["source"]                 = "ai"
    result["year_boundary_detected"] = year_boundary
    result["year_value"]             = year_value
    result["year_start"]             = year_start
    result["date_axis_row"]          = date_axis.get("row", 0)
    result["date_cols"]              = date_axis.get("cols", [])
    result["data_start_row"]         = grid.get("data_start_row", 1)
    result["resolved_dates"] = {
        str(k): v for k, v in
        session.get("_trusted_resolved_dates", {}).get(sheet_name, {}).items()
    }
    session["date_config"][sheet_name] = result
    _finalize_date_config(session, sheet_name, session["date_config"][sheet_name])


async def stage_date_config(session_id: str):
    """Stage 4 - date config. Pure Python, AI only if year ambiguous.
    Sheets are independent, so they run concurrently rather than one
    after another - see _date_config_one_sheet."""
    session = _sessions[session_id]
    session["stage"]  = "dating"
    session["status"] = "running"

    await asyncio.gather(*[
        _date_config_one_sheet(session, sheet_name)
        for sheet_name in session["qualified_sheets"]
    ])

    await stage_multisheet(session_id)


async def stage_multisheet(session_id: str):
    """Stage 5 - multi-sheet flag. Pure Python."""
    session   = _sessions[session_id]
    qualified = session["qualified_sheets"]
    multiple  = len(qualified) > 1
    session["flags"]["multiple_sales_sheets_detected"] = multiple
    if multiple:
        session["flags"]["multiple_sheets_note"] = (
            f"{len(qualified)} qualified sheets detected. "
            "Next stage should compare actual data to determine combine vs dedup."
        )
    await stage_write_audit(session_id)


async def stage_write_audit(session_id: str):
    """Stage 6 - query retailer config, then write to file_audit."""
    session       = _sessions[session_id]
    session["stage"]  = "writing"
    session["status"] = "running"
    file_audit_id = session.get("file_audit_id")
    retailer      = session.get("retailer")

    if not file_audit_id:
        await stage_assemble(session_id)
        return

    file_set_size = 1
    if retailer:
        try:
            sql  = build_query_retailer_config_sql(retailer)
            rows = await call_postgres(sql)
            if rows and rows[0].get("file_set_size") is not None:
                file_set_size = int(rows[0]["file_set_size"] or 1)
            else:
                file_set_size = None  # no active config
        except Exception as e:
            session["errors"].append(f"Failed to query retailer config: {e}")

    await stage_finalize_audit(session_id, file_set_size=file_set_size)


async def stage_finalize_audit(session_id: str, file_set_size: int = 1):
    """Write final status to file_audit."""
    session       = _sessions[session_id]
    file_audit_id = session.get("file_audit_id")
    retailer      = session.get("retailer")

    result = {
        "filename":         session.get("filename"),
        "file_hash":        session.get("file_hash"),
        "retailer":         retailer,
        "file_audit_id":    file_audit_id,
        "qualified_sheets": session.get("qualified_sheets", []),
        "qualify_results":  session.get("qualify_results", {}),
        "grid":             session.get("grid", {}),
        "column_mapping":   session.get("column_mapping", {}),
        "date_config":      session.get("date_config", {}),
        "flags":            session.get("flags", {}),
        "errors":           session.get("errors", []),
    }

    file_set_key = build_file_set_key(retailer, session.get("date_config", {}), session.get("grid", {})) if retailer else None

    # Only trust Postgres-confirmed retailer identification for automated ingestion
    # AI-confirmed retailer requires human review to prevent garbage data
    retailer_id_flag = session.get("flags", {}).get("retailer_identification", "")
    retailer_confirmed = retailer_id_flag.startswith("confirmed:")  # Postgres match only

    if not retailer:                audit_status = "pending_review"
    elif file_set_size is None:     audit_status = "pending_review"
    elif not retailer_confirmed:    audit_status = "pending_review"  # AI-identified - needs human sign-off
    elif file_set_size > 1:         audit_status = "pending_set"
    else:                           audit_status = "discovery_complete"

    sql = build_update_file_audit_full_sql(
        file_audit_id, result, retailer, audit_status, file_set_key,
        file_hash=session.get("file_hash"), filename=session.get("filename"),
        resolved_dates=session.get("_final_resolved_dates", {}),
    )
    try:
        await call_postgres(sql)
        # Write retailer config if retailer identified
        if retailer and file_audit_id:
            discovery_result = {
                "qualified_sheets": session.get("qualified_sheets", []),
                "column_mapping":   session.get("column_mapping", {}),
                "date_config":      session.get("date_config", {}),
                "flags":            session.get("flags", {}),
            }
            config_sql = build_insert_retailer_config_sql(retailer)
            await call_postgres(config_sql)
    except Exception as e:
        # Capture the exception TYPE, not just its string - "payload string
        # too long" means something very different from an asyncpg driver
        # exception versus a plain Python ValueError, and the bare message
        # alone doesn't distinguish them. Also measure the ACTUAL SQL size
        # that was sent, not an estimate - real evidence, not a guess, for
        # whoever looks at this next.
        sql_size_bytes = len(sql.encode("utf-8"))
        detailed = f"Failed to write audit/config: [{type(e).__name__}] {e} (SQL was {sql_size_bytes} bytes)"
        session["errors"].append(detailed)

        # If the full write failed, the row is otherwise left exactly as it
        # was before this stage ran (usually still 'analyzing') - silent
        # and invisible until the sweep eventually catches it, minutes
        # later, with none of this detail. Attempt a minimal, small
        # fallback write instead: just status + a diagnostic note, no large
        # JSONB payload - so the row becomes visibly failed immediately,
        # with the real reason, rather than sitting stuck.
        try:
            await call_postgres(build_sweep_fail_row_sql(file_audit_id, detailed))
        except Exception:
            pass  # if even the minimal write fails, the sweep is still the last resort

    await stage_assemble(session_id)


async def stage_assemble(session_id: str):
    """Stage 7 - assemble final result."""
    session = _sessions[session_id]
    session["stage"]  = "complete"
    session["status"] = "complete"
    session["result"] = {
        "filename":         session["filename"],
        "file_hash":        session["file_hash"],
        "retailer":         session.get("retailer"),
        "file_audit_id":    session.get("file_audit_id"),
        "qualified_sheets": session["qualified_sheets"],
        "qualify_results":  session["qualify_results"],
        "grid":             session["grid"],
        "column_mapping":   session["column_mapping"],
        "date_config":      session["date_config"],
        "flags":            session["flags"],
        "errors":           session.get("errors", []),
    }




# ---------------------------------------------
# ANALYZE ENDPOINT
# ---------------------------------------------

def file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

async def handle_discovery_file_binary(session_id: str, data: bytes, filename: str):
    """Called when file binary arrives from n8n for a discovery session."""
    if session_id not in _sessions:
        return

    session = _sessions[session_id]

    # Clean filename immediately — decode URL encoding and strip MinIO random suffix
    from urllib.parse import unquote
    filename = unquote(filename or "")
    filename = re.sub(r'_[A-Za-z0-9]{5}(\.[^.]+)$', r'\1', filename)

    if not filename.lower().endswith((".xlsx", ".xls")):
        file_audit_id = session.get("file_audit_id")
        if file_audit_id:
            try:
                sql = build_update_file_audit_full_sql(
                    file_audit_id,
                    {"status": "failed", "error": f"'{filename}' is not an Excel file"},
                    None, "failed", None,
                    filename=filename,
                )
                await call_postgres(sql)
            except Exception:
                pass
        session["stage"]  = "complete"
        session["status"] = "complete"
        session["result"] = {"error": f"'{filename}' is not an Excel file"}
        return

    # Immediate filename check - internal D365 inventory exports never need
    # sales-data qualification, so route them before opening the workbook at all.
    known_type = detect_known_file_type(filename, [])
    if known_type:
        file_audit_id = session.get("file_audit_id")
        if not session.get("file_hash"):
            session["file_hash"] = file_hash(data)
        if file_audit_id:
            try:
                sql = build_update_file_audit_full_sql(
                    file_audit_id,
                    {"status": "routed_by_filename"},
                    None, known_type, None,
                    file_hash=session.get("file_hash"),
                    filename=filename,
                )
                await call_postgres(sql)
            except Exception as e:
                session.setdefault("errors", []).append(f"Failed to write {known_type} status: {e}")
        session["stage"]  = "complete"
        session["status"] = "complete"
        session["result"] = {"status": known_type}
        return

    MAX_UPLOAD_BYTES = 50 * 1024 * 1024
    if len(data) > MAX_UPLOAD_BYTES:
        session["stage"]  = "complete"
        session["status"] = "complete"
        session["result"] = {"error": "File too large - max 50MB"}
        return

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        file_audit_id = session.get("file_audit_id")
        if file_audit_id:
            try:
                sql = build_update_file_audit_full_sql(
                    file_audit_id,
                    {"status": "failed", "error": f"Cannot open {filename}: {e}"},
                    None, "failed", None,
                    filename=filename,
                )
                await call_postgres(sql)
            except Exception:
                pass
        session["stage"]  = "complete"
        session["status"] = "complete"
        session["result"] = {"error": f"Cannot open {filename}: {e}"}
        return

    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets[sheet_name] = list(ws.iter_rows(values_only=True))

    session["_sheets"]   = sheets
    session["_raw_data"] = data
    session["filename"]  = filename
    if not session.get("file_hash"):
        session["file_hash"] = file_hash(data)

    # Dedup check - skip if this hash belongs to a different file_audit row
    try:
        dedup_rows = await call_postgres(build_dedup_check_sql(session["file_hash"]))
        if dedup_rows and str(dedup_rows[0].get("id")) != str(session.get("file_audit_id")):
            existing_id     = dedup_rows[0].get("id")
            existing_status = dedup_rows[0].get("status")
            session["stage"]  = "complete"
            session["status"] = "complete"
            session["result"] = {
                "skipped":   True,
                "reason":    f"File already processed - existing row {existing_id} has status '{existing_status}'",
                "file_hash": session["file_hash"],
            }
            return
    except Exception as e:
        session["errors"].append(f"Dedup check failed: {e}")

    await stage_qualify(session_id)


@router.post("/analyze")
async def analyze_from_audit(request_body: dict, background_tasks: BackgroundTasks = BackgroundTasks()):
    """
    Trigger discovery from a file_audit_id - file is fetched from MinIO via n8n.
    Used by the polling workflow for received files.
    """
    file_audit_id = request_body.get("file_audit_id")
    if not file_audit_id:
        raise HTTPException(status_code=400, detail="file_audit_id required")

    try:
        safe_id = _validate_uuid(file_audit_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid file_audit_id format")

    # Fetch audit row to get filename and check dedup
    try:
        rows = await call_postgres(build_fetch_audit_row_sql(safe_id))
        if not rows:
            raise HTTPException(status_code=404, detail=f"file_audit row {safe_id} not found")
        audit_row = rows[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch audit row: {e}")

    from urllib.parse import unquote
    filename  = unquote(audit_row.get("filename", "unknown.xlsx") or "unknown.xlsx")
    filename  = re.sub(r'_[A-Za-z0-9]{5}(\.[^.]+)$', r'\1', filename)
    file_hash_val = audit_row.get("file_hash")

    # Atomically claim this row - flips received -> analyzing.
    # If another concurrent poll already claimed it, bail out cleanly
    # instead of running discovery twice on the same file.
    try:
        claim_rows = await call_postgres(build_claim_file_audit_sql(safe_id))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to claim audit row: {e}")

    if not claim_rows:
        return JSONResponse(content={
            "status": "skipped",
            "reason": f"file_audit row {safe_id} is no longer 'received' - already claimed or processed",
        })

    session_id = str(uuid.uuid4())
    _sessions[session_id] = {
        "stage":            "accepted",
        "status":           "running",
        "filename":         filename,
        "file_hash":        file_hash_val,
        "file_audit_id":    safe_id,
        "retailer":         None,
        "qualified_sheets": [],
        "qualify_results":  {},
        "grid":             {},
        "postgres_results": {},
        "column_mapping":   {},
        "date_config":      {},
        "flags":            {},
        "errors":           [],
        "result":           None,
        "_sheets":          None,   # populated when file arrives
        "_raw_data":        None,
        "_pending_jobs":    set(),
        "created_at":       time.time(),
    }

    # Fire fetch-file webhook - file binary will arrive at /file/{job_id}
    # But discovery needs the binary before it can start - use a flag to wait
    file_job_id = str(uuid.uuid4())
    _jobs[file_job_id] = {
        "session_id": session_id,
        "stage":      "fetch_file_binary",
        "audit_id":   safe_id,
        "pipeline":   "discovery",
        "created_at": time.time(),
    }
    err = await fire_fetch_file_webhook(file_job_id, safe_id)
    if err:
        _sessions.pop(session_id, None)
        _jobs.pop(file_job_id, None)
        raise HTTPException(status_code=500, detail=f"Failed to fetch file: {err}")

    _sessions[session_id]["_pending_jobs"].add(file_job_id)

    async def _wait_for_file():
        """
        Poll for completion. The actual file binary is delivered to
        POST /file/{job_id} (in ingestion.py), which invokes
        handle_discovery_file_binary directly as a background task.
        This loop just waits for that task to finish - it never overwrites
        a session that has progressed past the initial 'accepted' stage,
        since that would mean real processing is underway (e.g. AI column
        classification across multiple sheets, which can legitimately take
        a while) and must not be stomped by an artificial timeout.
        """
        for _ in range(300):  # wait up to 5 minutes
            await asyncio.sleep(1)
            session = _sessions.get(session_id)
            if not session:
                return
            if session.get("status") in ("complete", "failed"):
                return
        # Timed out - only treat as a real failure if the file binary itself
        # never arrived (session never progressed past "accepted"/"running"
        # with no sheets loaded). Otherwise leave the session alone; it is
        # still legitimately processing and will complete on its own.
        session = _sessions.get(session_id)
        if session and session.get("_sheets") is None:
            session["stage"]  = "complete"
            session["status"] = "failed"
            session["result"] = {"error": "Timed out waiting for file binary from MinIO"}

    background_tasks.add_task(_wait_for_file)

    return JSONResponse(content={
        "status":     "accepted",
        "session_id": session_id,
        "poll_url":   f"/analyze/status/{session_id}",
    })



@router.get("/analyze/status/{session_id}")
async def analyze_status(session_id: str):
    if session_id not in _sessions:
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    session = _sessions[session_id]
    return JSONResponse(content={
        "session_id": session_id,
        "stage":      session.get("stage", "unknown"),
        "status":     session.get("status", "unknown"),
        "filename":   session["filename"],
        "result":     session["result"],
    })


@router.get("/health")
def health():
    return {"status": "ok", "version": "3.2.0"}

"""
ingestion.py — Ingestion pipeline: fetch audit rows, process files, write sales and inventory.
"""
from fastapi import APIRouter, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from typing import List, Optional
import openpyxl
import io
import re
import uuid
import asyncio
import urllib.request
import json
from datetime import datetime, date, timedelta
import time

from shared import (
    _sessions, _jobs,
    call_postgres, fire_fetch_file_webhook,
    normalize_to_saturday,
    _validate_uuid, _validate_date, _sql_escape,
    build_fetch_audit_row_sql,
)

router = APIRouter()

# ─────────────────────────────────────────────
# DATE HELPERS
# ─────────────────────────────────────────────
# There is no date-resolution logic in this file anymore. Discovery computes
# and persists the complete col_idx -> date mapping (date_config[sheet]
# ["resolved_dates"]) and the inventory as-of date (date_config[sheet]
# ["inventory_as_of_date"]) for every qualified sheet - see discovery.py's
# _finalize_date_config / build_date_map (in shared.py). Ingestion reads
# those values directly; it never re-derives a date from header text,
# never touches the pattern table, never needs to know what row the date
# axis lives in. If a date is missing here, that's a discovery bug to fix,
# not something ingestion should independently try to work around -
# exactly the shape of bug (a stale pattern cache, a wrong header row) that
# silently produced zero-row "successful" ingestions when this logic used
# to live here, duplicated and independently re-derived.


# ─────────────────────────────────────────────
# SQL BUILDERS
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# SQL HELPERS
# ─────────────────────────────────────────────

def build_fetch_retailer_config_sql(retailer: str) -> str:
    retailer_safe = retailer.replace("'", "''")
    return f"""
SELECT id, retailer, file_set_size, qualified_sheets, column_mapping, date_config, flags
FROM retailer_configs
WHERE UPPER(retailer) = UPPER('{retailer_safe}')
  AND status = 'active'
LIMIT 1
""".strip()


def build_fetch_sku_map_sql(retailer: str, retailer_skus: list) -> str:
    retailer_safe = retailer.replace("'", "''")
    quoted = ", ".join("'" + str(s).replace("'", "''") + "'" for s in retailer_skus if s)
    if not quoted:
        quoted = "''"
    return f"""
SELECT retailer_sku, supplier_sku, base_model, base_variant, description
FROM retailer_sku_map
WHERE UPPER(retailer) = UPPER('{retailer_safe}')
  AND UPPER(retailer_sku) = ANY(ARRAY[{quoted}]::text[])
  AND active = true
""".strip()



def build_lookup_supplier_skus_sql(retailer: str, retailer_skus: list) -> str:
    """Look up supplier SKUs from retailer_sku_map for a list of retailer SKUs."""
    retailer_safe = retailer.replace("'", "''")
    quoted = ", ".join("'" + str(s).replace("'", "''") + "'" for s in retailer_skus if s)
    if not quoted:
        quoted = "''"
    return f"""
SELECT retailer_sku, supplier_sku, base_model, base_variant
FROM retailer_sku_map
WHERE UPPER(retailer) = UPPER('{retailer_safe}')
  AND UPPER(retailer_sku) = ANY(ARRAY[{quoted}]::text[])
  AND active = true
  AND supplier_sku IS NOT NULL
""".strip()


def build_upsert_sales_sql(retailer: str, rows: list, file_audit_id: str) -> str:
    """Build a multi-row upsert into unified weekly_sales table."""
    safe_retailer = _sql_escape(retailer)
    safe_fid      = _validate_uuid(file_audit_id)
    values        = []
    for row in rows:
        retailer_sku = _sql_escape(row["retailer_sku"])
        supplier_sku = f"'{_sql_escape(row['supplier_sku'])}'" if row.get("supplier_sku") else "NULL"
        week_ending  = _validate_date(row["week_ending"])
        units_sold   = int(row["units_sold"])
        values.append(
            f"('{safe_retailer}', '{retailer_sku}', {supplier_sku}, '{week_ending}', {units_sold}, '{safe_fid}')"
        )

    values_str = ",\n".join(values)
    return f"""
INSERT INTO weekly_sales (retailer, retailer_sku, supplier_sku, week_ending, units_sold, file_audit_id)
VALUES {values_str}
ON CONFLICT (retailer, retailer_sku, week_ending)
DO UPDATE SET
    units_sold    = EXCLUDED.units_sold,
    supplier_sku  = EXCLUDED.supplier_sku,
    file_audit_id = EXCLUDED.file_audit_id,
    ingested_at   = now()
WHERE weekly_sales.locked = false
""".strip()


def build_upsert_inventory_sql(retailer: str, rows: list, file_audit_id: str, as_of_date: str) -> str:
    """Build a multi-row upsert for inventory snapshot."""
    safe_retailer = _sql_escape(retailer)
    safe_fid      = _validate_uuid(file_audit_id)
    safe_date     = _validate_date(as_of_date)
    values        = []
    for row in rows:
        retailer_sku = _sql_escape(row["retailer_sku"])
        on_hand      = int(row.get("on_hand_qty") or 0)
        open_order   = int(row.get("open_order_qty") or 0)
        values.append(f"('{safe_retailer}', '{retailer_sku}', {on_hand}, {open_order}, '{safe_date}', '{safe_fid}')")

    values_str = ",\n".join(values)
    return f"""
INSERT INTO retailer_inventory (retailer, retailer_sku, on_hand_qty, open_order_qty, as_of_date, file_audit_id)
VALUES {values_str}
ON CONFLICT (retailer, retailer_sku)
DO UPDATE SET
    on_hand_qty    = EXCLUDED.on_hand_qty,
    open_order_qty = EXCLUDED.open_order_qty,
    as_of_date     = EXCLUDED.as_of_date,
    file_audit_id  = EXCLUDED.file_audit_id,
    updated_at     = now()
WHERE retailer_inventory.as_of_date <= EXCLUDED.as_of_date
""".strip()


def build_update_audit_status_sql(file_audit_id: str, status: str,
                                   unresolved_skus: list = None) -> str:
    safe_id         = _validate_uuid(file_audit_id)
    safe_status     = _sql_escape(status)
    unresolved_json = _sql_escape(json.dumps(unresolved_skus or []))
    return f"""
UPDATE file_audit
SET status          = '{safe_status}',
    unresolved_skus = '{unresolved_json}'::jsonb,
    updated_at      = now()
WHERE id = '{safe_id}'
""".strip()


async def mark_session_failed(session_id: str, error_message: str, extra_errors: list = None):
    """
    The single, consistent way to fail an ingestion session. Updates BOTH
    the in-memory session state (for anything actively polling this run)
    AND file_audit's status in Postgres - the persistent record.

    Before this existed, 6 separate failure points across this file only
    updated the in-memory session dict. That dict is ephemeral - per
    shared.py's cleanup_stale_jobs, it's evicted SESSION_GRACE_SECONDS after
    completion/failure. Meanwhile file_audit.status was only ever written
    on the SUCCESS path (stage_finalize, "ingested"/"ingested_partial") -
    never on failure. Net effect: a file that failed partway through stayed
    stuck at whatever status it had before failing (typically 'ingesting')
    forever, with zero persistent record anything went wrong, once the
    in-memory session was evicted.

    Best-effort on the Postgres write: if IT fails too, don't let that mask
    or crash reporting of the ORIGINAL error - the in-memory session update
    still happens regardless.
    """
    session = _sessions[session_id]
    session["stage"]  = "failed"
    session["status"] = "failed"
    session["result"] = {
        "error": error_message,
        "errors": extra_errors if extra_errors is not None else session.get("errors", []),
    }

    for audit_id in session.get("file_audit_ids", []) or []:
        try:
            await call_postgres(build_update_audit_status_sql(
                audit_id, "failed", session.get("unresolved_skus", [])
            ))
        except Exception:
            pass


# ─────────────────────────────────────────────
# CORE INGESTION LOGIC
# ─────────────────────────────────────────────

def download_from_url(url: str) -> Optional[bytes]:
    """Download file from a URL."""
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        with urllib.request.urlopen(req, timeout=30) as resp:
            return resp.read()
    except Exception as e:
        return None


def extract_sales_and_inventory(
    wb: openpyxl.Workbook,
    discovery_result: dict,
    retailer: str,
    file_audit_id: str,
    resolved_dates: dict | None = None,
) -> dict:
    """
    Extract sales rows and inventory rows from workbook using discovery result.
    Returns {sales: [...], inventory: [...], unresolved_skus: [...]}

    resolved_dates lives in its own top-level file_audit column now, not
    nested inside discovery_result/date_config - kept in exactly one place,
    not duplicated between a JSON sub-field and a dedicated column.
    """
    column_mapping   = discovery_result.get("column_mapping", {})
    date_config      = discovery_result.get("date_config", {})
    qualified_sheets = discovery_result.get("qualified_sheets", [])
    resolved_dates   = resolved_dates or {}

    # Accumulate sales by (retailer_sku, week_ending) for rollup across sheets
    sales_map      = {}  # (retailer_sku, week_ending) -> units_sold
    sku_supplier   = {}  # retailer_sku -> supplier_sku
    inventory_map  = {}  # retailer_sku -> {on_hand_qty, open_order_qty}
    unresolved     = set()
    inv_as_of_date = None  # most recent inventory snapshot date from column headers

    for sheet_name in qualified_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws      = wb[sheet_name]
        rows    = list(ws.iter_rows(values_only=True))
        mapping = column_mapping.get(sheet_name, [])
        dc      = date_config.get(sheet_name, {})

        if not mapping or not rows:
            continue

        # Find column roles
        retailer_sku_col  = None
        supplier_sku_col  = None
        description_col   = None
        inventory_col     = None
        open_order_col    = None
        date_cols         = []

        for col_info in mapping:
            classification = col_info.get("classification")
            col_idx        = col_info.get("col")
            if classification == "retailer_sku":
                retailer_sku_col = col_idx
            elif classification == "supplier_sku":
                supplier_sku_col = col_idx
            elif classification == "description":
                description_col = col_idx
            elif classification == "inventory":
                inventory_col = col_idx
            elif classification == "open_order":
                open_order_col = col_idx

        # Discovery persists everything date-related now - this reads it,
        # it never figures anything out. See discovery.py's
        # _finalize_date_config for where these two fields get computed.
        date_col_idxs = dc.get("date_cols", [])
        data_start    = dc.get("data_start_row", 1)

        if not date_col_idxs or retailer_sku_col is None:
            continue

        # Inventory as-of date: already computed by discovery, per sheet.
        # Keep the most recent one across sheets, same as before.
        if inventory_col is not None:
            sheet_inv_date = dc.get("inventory_as_of_date")
            if sheet_inv_date and (inv_as_of_date is None or sheet_inv_date > inv_as_of_date):
                inv_as_of_date = sheet_inv_date

        # Sales date map: a plain read of what discovery already computed
        # and verified - a col_idx string key -> ISO date string dict, one
        # entry per real date column, sourced from its own dedicated
        # file_audit.resolved_dates column. No recomputation, no header
        # lookup, no pattern matching, no year-tracking state.
        date_map = {}
        for k, v in resolved_dates.get(sheet_name, {}).items():
            try:
                date_map[int(k)] = date.fromisoformat(v)
            except (ValueError, TypeError):
                continue

        if not date_map:
            continue

        # Process data rows
        for row in rows[data_start:]:
            if retailer_sku_col >= len(row) or row[retailer_sku_col] is None:
                continue

            rsku = str(row[retailer_sku_col]).strip().strip("\r\n")
            if not rsku or rsku.lower() in ("total", "grand total", "subtotal", ""):
                continue

            # Supplier SKU
            ssku = None
            if supplier_sku_col is not None and supplier_sku_col < len(row):
                ssku = str(row[supplier_sku_col]).strip().strip('\r\n') if row[supplier_sku_col] else None
            if ssku:
                sku_supplier[rsku] = ssku

            # Inventory
            if inventory_col is not None and inventory_col < len(row):
                v = row[inventory_col]
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    inv = inventory_map.get(rsku, {"on_hand_qty": 0, "open_order_qty": 0})
                    inv["on_hand_qty"] = int(v)
                    inventory_map[rsku] = inv

            # Open orders
            if open_order_col is not None and open_order_col < len(row):
                v = row[open_order_col]
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    inv = inventory_map.get(rsku, {"on_hand_qty": 0, "open_order_qty": 0})
                    inv["open_order_qty"] = int(v)
                    inventory_map[rsku] = inv

            # Sales — each date column
            for col_idx, week_end in date_map.items():
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None or isinstance(val, bool):
                    continue
                if not isinstance(val, (int, float)):
                    continue
                units = int(val)
                if units == 0:
                    continue

                key = (rsku, str(week_end))
                sales_map[key] = sales_map.get(key, 0) + units

    # Build output rows
    sales_rows = []
    for (rsku, week_end), units in sales_map.items():
        sales_rows.append({
            "retailer_sku": rsku,
            "supplier_sku": sku_supplier.get(rsku),
            "week_ending":  week_end,
            "units_sold":   units,
        })
        if not sku_supplier.get(rsku):
            unresolved.add(rsku)

    inventory_rows = []
    for rsku, inv in inventory_map.items():
        inventory_rows.append({
            "retailer_sku":  rsku,
            "on_hand_qty":   inv.get("on_hand_qty", 0),
            "open_order_qty": inv.get("open_order_qty", 0),
        })

    return {
        "sales":           sales_rows,
        "inventory":       inventory_rows,
        "unresolved_skus": sorted(unresolved),
        "inv_as_of_date":  inv_as_of_date,
    }


# ─────────────────────────────────────────────
# PIPELINE STAGES
# ─────────────────────────────────────────────


# fetch_file_binary is the only remaining webhook job in ingestion
# If it times out, attempt to process with whatever audit rows we have
async def _file_fetch_timeout_handler(sid: str, stage: str):
    if stage == "fetch_file_binary":
        session = _sessions.get(sid)
        if session and session.get("_audit_rows") and session.get("_file_bytes"):
            asyncio.create_task(stage_process_files(sid))
        elif session:
            await mark_session_failed(sid, "File fetch timed out")


class IngestRequest(BaseModel):
    file_audit_ids: List[str]


@router.post("/ingest")
async def ingest(request: IngestRequest, background_tasks: BackgroundTasks):
    if not request.file_audit_ids:
        raise HTTPException(status_code=400, detail="At least one file_audit_id required")

    session_id = str(uuid.uuid4())
    _sessions[session_id] = {
        "stage":           "accepted",
        "status":          "running",
        "file_audit_ids":  request.file_audit_ids,
        "retailer":        None,
        "sales_rows":      [],
        "inventory_rows":  [],
        "unresolved_skus": [],
        "errors":          [],
        "result":          None,
        "_pending_jobs":   set(),  # only for fetch_file_binary jobs
        "_audit_rows":     {},
        "created_at":      time.time(),
    }

    background_tasks.add_task(stage_fetch_audit_rows, session_id)

    return JSONResponse(content={
        "status":     "accepted",
        "session_id": session_id,
        "poll_url":   f"/ingest/status/{session_id}",
    })


async def stage_fetch_audit_rows(session_id: str):
    """Fetch all file_audit rows from Postgres and request file binaries from n8n."""
    session = _sessions[session_id]
    session["stage"]  = "fetching"
    session["status"] = "running"

    # Mark all files as ingesting immediately — Python owns the status from here
    for audit_id in session["file_audit_ids"]:
        try:
            await call_postgres(f"""
UPDATE file_audit SET status = 'ingesting', updated_at = now()
WHERE id = '{_validate_uuid(audit_id)}'
""".strip())
        except Exception as e:
            session["errors"].append(f"Failed to mark ingesting for {audit_id}: {e}")

    for audit_id in session["file_audit_ids"]:
        # Fetch audit row directly from Postgres
        try:
            sql  = build_fetch_audit_row_sql(audit_id)
            rows = await call_postgres(sql)
            if rows:
                row = rows[0]
                dr  = row.get("discovery_result")
                if isinstance(dr, str):
                    try:
                        dr = json.loads(dr)
                    except (json.JSONDecodeError, ValueError):
                        dr = {}
                row["discovery_result"] = dr
                session["_audit_rows"][audit_id] = row
            else:
                session["errors"].append(f"No audit row found for {audit_id}")
        except Exception as e:
            session["errors"].append(f"Failed to fetch audit row {audit_id}: {e}")

        # File binary still comes from n8n (MinIO fetch)
        file_job_id = str(uuid.uuid4())
        _jobs[file_job_id] = {
            "session_id": session_id,
            "stage":      "fetch_file_binary",
            "audit_id":   audit_id,
            "created_at": time.time(),
            "pipeline":   "ingestion",
        }
        err = await fire_fetch_file_webhook(file_job_id, audit_id)
        if err:
            session["errors"].append(f"Failed to fetch file binary for {audit_id}: {err}")
        else:
            session["_pending_jobs"].add(file_job_id)

    # If all file fetches failed, fail the session
    if not session["_pending_jobs"] and not session.get("_file_bytes"):
        await mark_session_failed(session_id, "Failed to fetch any file binaries")


async def stage_process_files(session_id: str):
    """Process file binaries — extract sales and inventory from all audit rows."""
    session = _sessions[session_id]
    session["stage"]  = "processing"
    session["status"] = "running"

    audit_rows = session.get("_audit_rows", {})
    if not audit_rows:
        await mark_session_failed(session_id, "No audit rows found")
        return

    retailers = set(r.get("retailer") for r in audit_rows.values() if r.get("retailer"))
    if not retailers:
        await mark_session_failed(session_id, "No retailer identified in audit rows")
        return

    retailer = retailers.pop()
    session["retailer"] = retailer

    all_sales     = {}
    all_inventory = {}
    all_unresolved = set()
    as_of_dates   = []
    file_bytes    = session.get("_file_bytes", {})

    for audit_id, audit_row in audit_rows.items():
        discovery_result = audit_row.get("discovery_result")
        resolved_dates   = audit_row.get("resolved_dates") or {}
        data             = file_bytes.get(audit_id)

        if not discovery_result:
            session["errors"].append(f"Missing discovery_result for {audit_id}")
            continue
        if not data:
            session["errors"].append(f"No file data available for {audit_id}")
            continue

        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            session["errors"].append(f"Failed to open workbook for {audit_id}: {e}")
            continue

        extracted = extract_sales_and_inventory(wb, discovery_result, retailer, audit_id, resolved_dates)

        for row in extracted["sales"]:
            key = (row["retailer_sku"], row["week_ending"])
            all_sales[key] = all_sales.get(key, 0) + row["units_sold"]
            if row.get("supplier_sku"):
                session.setdefault("_sku_supplier", {})[row["retailer_sku"]] = row["supplier_sku"]

        for row in extracted["inventory"]:
            all_inventory[row["retailer_sku"]] = row

        all_unresolved.update(extracted["unresolved_skus"])

        if extracted.get("inv_as_of_date"):
            as_of_dates.append(extracted["inv_as_of_date"])
        else:
            for row in extracted["sales"]:
                as_of_dates.append(row["week_ending"])

    sku_supplier = session.get("_sku_supplier", {})
    session["sales_rows"] = [
        {
            "retailer_sku": rsku,
            "supplier_sku": sku_supplier.get(rsku),
            "week_ending":  week_end,
            "units_sold":   units,
        }
        for (rsku, week_end), units in all_sales.items()
    ]
    session["inventory_rows"]  = list(all_inventory.values())
    session["unresolved_skus"] = sorted(all_unresolved)
    session["as_of_date"]      = max(as_of_dates) if as_of_dates else str(date.today())

    await stage_lookup_supplier_skus(session_id)


async def stage_lookup_supplier_skus(session_id: str):
    """Look up supplier SKUs from retailer_sku_map.
    retailer_sku_map is always the source of truth — any supplier_sku
    extracted from the file itself is overwritten by the mapped value.
    """
    session   = _sessions[session_id]
    retailer  = session["retailer"]
    session["stage"]  = "looking_up_skus"
    session["status"] = "running"

    sales_rows    = session.get("sales_rows", [])
    retailer_skus = list({row["retailer_sku"] for row in sales_rows})

    if retailer_skus:
        try:
            sql     = build_lookup_supplier_skus_sql(retailer, retailer_skus)
            rows    = await call_postgres(sql)
            sku_map = {r["retailer_sku"]: r for r in rows if r.get("supplier_sku")}
            for row in sales_rows:
                if row["retailer_sku"] in sku_map:
                    row["supplier_sku"] = sku_map[row["retailer_sku"]].get("supplier_sku")
            session["sales_rows"] = sales_rows
        except Exception as e:
            session["errors"].append(f"Failed to look up supplier SKUs: {e}")

    session["unresolved_skus"] = sorted({
        row["retailer_sku"] for row in sales_rows if not row.get("supplier_sku")
    })

    # stage_validate_schema used to sit here, checking for a per-retailer
    # convenience view ("{retailer}_weekly_sales") before allowing writes.
    # Removed: the actual write (build_upsert_sales_sql) always targets the
    # single shared weekly_sales table directly - retailer is just a column
    # value there - so the view's existence had no bearing on whether the
    # write would succeed. It was blocking valid data on the absence of an
    # optional reporting convenience the write path never used.
    await stage_write_sales(session_id)


async def stage_write_sales(session_id: str):
    """Write sales rows to retailer sales table in batches."""
    session   = _sessions[session_id]
    retailer  = session["retailer"]
    session["stage"]  = "writing_sales"
    session["status"] = "running"

    sales_rows = session.get("sales_rows", [])
    if not sales_rows:
        await stage_write_inventory(session_id)
        return

    audit_id   = session["file_audit_ids"][0]
    BATCH_SIZE = 500

    for i, batch in enumerate([sales_rows[i:i+BATCH_SIZE] for i in range(0, len(sales_rows), BATCH_SIZE)]):
        try:
            await call_postgres(build_upsert_sales_sql(retailer, batch, audit_id))
        except Exception as e:
            session["errors"].append(f"Failed to write sales batch {i}: {e}")

    await stage_write_inventory(session_id)


async def stage_write_inventory(session_id: str):
    """Write inventory snapshot in batches."""
    session   = _sessions[session_id]
    retailer  = session["retailer"]
    session["stage"]  = "writing_inventory"
    session["status"] = "running"

    inventory_rows = session.get("inventory_rows", [])
    if not inventory_rows:
        await stage_finalize(session_id)
        return

    audit_id   = session["file_audit_ids"][0]
    as_of_date = session.get("as_of_date", str(date.today()))
    BATCH_SIZE = 500

    for i, batch in enumerate([inventory_rows[i:i+BATCH_SIZE] for i in range(0, len(inventory_rows), BATCH_SIZE)]):
        try:
            await call_postgres(build_upsert_inventory_sql(retailer, batch, audit_id, as_of_date))
        except Exception as e:
            session["errors"].append(f"Failed to write inventory batch {i}: {e}")

    await stage_finalize(session_id)


async def stage_finalize(session_id: str):
    """Update file_audit status and refresh materialized forecast views."""
    session      = _sessions[session_id]
    unresolved   = session.get("unresolved_skus", [])
    audit_status = "ingested_partial" if unresolved else "ingested"

    for audit_id in session["file_audit_ids"]:
        try:
            await call_postgres(build_update_audit_status_sql(audit_id, audit_status, unresolved))
        except Exception as e:
            session["errors"].append(f"Failed to update audit status for {audit_id}: {e}")

    # Refresh materialized forecast views — non-fatal if they fail
    for view in ("mv_sku_seasonality", "mv_weekly_forecast"):
        try:
            await call_postgres(f"REFRESH MATERIALIZED VIEW CONCURRENTLY {view}")
        except Exception as e:
            session["errors"].append(f"View refresh failed for {view} (non-fatal): {e}")

    await stage_complete(session_id)


async def stage_complete(session_id: str):
    session = _sessions[session_id]
    session["stage"]  = "complete"
    session["status"] = "complete"
    session["result"] = {
        "retailer":        session.get("retailer"),
        "file_audit_ids":  session["file_audit_ids"],
        "sales_rows":      len(session.get("sales_rows", [])),
        "inventory_rows":  len(session.get("inventory_rows", [])),
        "unresolved_skus": session.get("unresolved_skus", []),
        "errors":          session.get("errors", []),
    }


# ─────────────────────────────────────────────
# STATUS ENDPOINT
# ─────────────────────────────────────────────



@router.post("/file/{job_id}")
async def file_upload_response(job_id: str, file: UploadFile = File(...), background_tasks: BackgroundTasks = BackgroundTasks()):
    """Receive binary file from n8n S3 webhook — handles both discovery and ingestion pipelines."""
    if job_id not in _jobs:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    job        = _jobs.pop(job_id)
    session_id = job["session_id"]
    pipeline   = job.get("pipeline", "ingestion")

    if session_id not in _sessions:
        return JSONResponse(content={"status": "ok", "note": "session already closed"})

    session = _sessions[session_id]
    session["_pending_jobs"].discard(job_id)
    data = await file.read()

    if pipeline == "discovery":
        # Route to discovery pipeline handler
        from discovery import handle_discovery_file_binary
        background_tasks.add_task(handle_discovery_file_binary, session_id, data, file.filename)
    else:
        # Ingestion pipeline
        audit_id = job["audit_id"]
        session.setdefault("_file_bytes", {})[audit_id] = data
        session.setdefault("_filenames", {})[audit_id] = file.filename

        expected = len(session["file_audit_ids"])
        if (not session["_pending_jobs"]
                and len(session.get("_audit_rows", {})) >= expected
                and len(session.get("_file_bytes", {})) >= expected):
            background_tasks.add_task(stage_process_files, session_id)

    return JSONResponse(content={"status": "ok", "job_id": job_id})

@router.get("/ingest/status/{session_id}")
async def ingest_status(session_id: str):
    if session_id not in _sessions:
        raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    session = _sessions[session_id]
    return JSONResponse(content={
        "session_id": session_id,
        "stage":      session.get("stage"),
        "status":     session.get("status"),
        "result":     session.get("result"),
    })


@router.get("/ingest/health")
def health():
    return {"status": "ok", "version": "1.0.0"}
